import pandas as pd, sys, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

# ── 讀取資料 ──
df = pd.read_excel(r'C:\Users\admin\OneDrive\桌面\20240102-20260508進貨單價.xlsx', dtype=str)
df['品號'] = df['品    號'].fillna('').str.strip()
df['品名'] = df['品    名'].fillna('').str.strip()
df['規格'] = df['規    格'].fillna('').str.strip()
df['廠商簡稱'] = df['廠商簡稱'].fillna('').str.strip()
df['單位進價'] = pd.to_numeric(df['原幣單位進價'], errors='coerce')
df['日期'] = pd.to_datetime(df['單據日期'], errors='coerce')
df['單位'] = df['單位'].fillna('').str.strip()
df = df[df['品號'] != ''].copy()

# ── 樣式常數 ──
C_DARK   = '1C3F6B'
C_GREEN  = '1A5C3A'
C_RED    = 'CC2200'
C_ORANGE = 'CC7700'
C_PURPLE = '5B2D8E'
C_TEAL   = '006D77'
C_LBLUE  = 'D6E8FB'
C_LGREEN = 'E8F5E9'
C_LRED   = 'FFF0F0'
C_LYELLOW= 'FFFDE7'
C_WHITE  = 'FFFFFF'
C_GRAY   = 'F5F5F5'
C_GOLD   = 'FFC300'

def set_border(cell, style='thin', color='CCCCCC'):
    s = Side(style=style, color=color)
    cell.border = Border(left=s, right=s, top=s, bottom=s)

def hd(ws, r, c, v, bg=C_DARK, fc='FFFFFF', sz=11, bold=True, align='center', wrap=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name='微軟正黑體', bold=bold, color=fc, size=sz)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    set_border(cell, 'medium', '888888')
    return cell

def dc(ws, r, c, v, bg=C_WHITE, fc='222222', sz=10, bold=False, align='left', fmt=None, italic=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name='微軟正黑體', bold=bold, size=sz, color=fc, italic=italic)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)
    set_border(cell, 'hair', 'DDDDDD')
    if fmt: cell.number_format = fmt
    return cell

# ── 依品號大類 + 廠商取最新進價統計 ──
df_s = df.sort_values('日期')
def get_stats(sub):
    return sub.groupby(['品號','品名','規格','廠商簡稱','單位']).agg(
        最低進價=('單位進價','min'),
        最高進價=('單位進價','max'),
        平均進價=('單位進價','mean'),
        最新進價=('單位進價','last'),
        最新日期=('日期','last'),
        進貨次數=('單位進價','count')
    ).reset_index()

# 各供應商按品類整理最新/最低/最高進價（pivot格式）
def make_pivot(sub_df):
    """回傳每個品號的多廠商比較寬表"""
    g = sub_df.sort_values('日期').groupby(['品號','品名','規格','單位','廠商簡稱']).agg(
        最新進價=('單位進價','last'),
        最低進價=('單位進價','min'),
        最高進價=('單位進價','max'),
        進貨次數=('單位進價','count'),
        最新日期=('日期','last'),
    ).reset_index()
    return g

# ── 工作表設定 ──
SHEETS = [
    ('總覽比價', None, C_DARK),
    ('BPP 外層布', 'BPP', '1A5C3A'),
    ('BEL 耳繩',  'BEL', '5B2D8E'),
    ('BMB 熔噴布','BMB', '006D77'),
    ('BES 親膚布','BES', 'B45309'),
    ('FBOX 外箱', 'FBOX','CC2200'),
    ('BPA 包裝',  'BPA', '1C3F6B'),
    ('BAC 活性碳','BAC', '36454F'),
    ('BJF 接合布','BJF', '36454F'),
    ('BNW 鼻樑條','BNW', '36454F'),
    ('其他材料',  None,  '36454F'),
]
OTHER_CATS = ['FPEB','FKOP','BPE','BQC','FDM','CARD','FSH','FPA','FDRX']

wb = Workbook()
wb.remove(wb.active)

# ────────────────────────────────────────────────
# 通用函數：寫一張比價工作表
# ────────────────────────────────────────────────
def write_sheet(ws, sub_df, title, hdr_color):
    if len(sub_df) == 0:
        ws.cell(1,1,'（無資料）')
        return

    g = make_pivot(sub_df)
    suppliers = sorted(g['廠商簡稱'].unique().tolist())
    n_sup = len(suppliers)

    # ── 標題 ──
    total_cols = 4 + n_sup * 4 + 3   # 品號品名規格單位 + 每廠商4欄 + 最優/價差/備註
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws.cell(row=1, column=1,
        value=f'{title}  進貨比價表  |  資料期間: 2024-01 ~ 2026-05  |  共 {g["品號"].nunique()} 品項')
    t.font = Font(name='微軟正黑體', bold=True, size=13, color='FFFFFF')
    t.fill = PatternFill('solid', start_color=hdr_color)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # ── 廠商分組標題 ──
    col = 5
    for sup in suppliers:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+3)
        sc = ws.cell(row=2, column=col, value=sup)
        sc.font = Font(name='微軟正黑體', bold=True, size=11, color='FFFFFF')
        sc.fill = PatternFill('solid', start_color=hdr_color)
        sc.alignment = Alignment(horizontal='center', vertical='center')
        set_border(sc, 'medium', '888888')
        col += 4
    ws.row_dimensions[2].height = 22

    # ── 欄標題列 ──
    base_hdrs = ['品號','品名','規格','單位']
    for ci, h in enumerate(base_hdrs, 1):
        hd(ws, 3, ci, h, bg=hdr_color)
    col = 5
    for sup in suppliers:
        hd(ws, 3, col,   '最新進價', bg=hdr_color)
        hd(ws, 3, col+1, '最低進價', bg=hdr_color)
        hd(ws, 3, col+2, '最高進價', bg=hdr_color)
        hd(ws, 3, col+3, '進貨次數', bg=hdr_color)
        col += 4
    hd(ws, 3, col,   '最優廠商', bg=C_GREEN)
    hd(ws, 3, col+1, '最低價差%', bg=C_GREEN)
    hd(ws, 3, col+2, '備註', bg=C_DARK)
    ws.row_dimensions[3].height = 24

    # ── 資料 ──
    items = g.groupby(['品號','品名','規格','單位'])
    row = 4
    for (pno, pname, spec, unit), grp in items:
        bg = C_LBLUE if row % 2 == 0 else C_WHITE
        dc(ws, row, 1, pno,   bg=bg, sz=10)
        dc(ws, row, 2, pname, bg=bg, sz=10, bold=True)
        dc(ws, row, 3, spec,  bg=bg, sz=9, fc='555555')
        dc(ws, row, 4, unit,  bg=bg, align='center')

        prices_by_sup = {}
        col = 5
        for sup in suppliers:
            r = grp[grp['廠商簡稱']==sup]
            if len(r) > 0:
                newest = r.iloc[0]['最新進價']
                lo     = r.iloc[0]['最低進價']
                hi     = r.iloc[0]['最高進價']
                cnt    = r.iloc[0]['進貨次數']
                prices_by_sup[sup] = newest
                # 若最新比最低高>5%，標橘色警示
                hilight = newest > lo * 1.05
                dc(ws, row, col,   newest, bg=C_LYELLOW if hilight else bg, fc=C_RED if hilight else '111111', bold=hilight, align='right', fmt='#,##0.0000')
                dc(ws, row, col+1, lo,     bg=bg, align='right', fmt='#,##0.0000', fc=C_GREEN)
                dc(ws, row, col+2, hi,     bg=bg, align='right', fmt='#,##0.0000', fc=C_RED if hi > lo else '111111')
                dc(ws, row, col+3, int(cnt), bg=bg, align='center', fc='444444')
            else:
                for ci in range(4):
                    dc(ws, row, col+ci, '—', bg=C_GRAY, align='center', fc='AAAAAA', italic=True)
            col += 4

        # 最優廠商分析
        valid = {s:p for s,p in prices_by_sup.items() if pd.notna(p) and p > 0}
        if len(valid) >= 2:
            best_sup = min(valid, key=valid.get)
            best_p   = valid[best_sup]
            worst_p  = max(valid.values())
            diff_pct = round((worst_p - best_p) / worst_p * 100, 1)
            dc(ws, row, col,   best_sup, bg=C_LGREEN, bold=True, fc=C_GREEN, align='center')
            dc(ws, row, col+1, f'{diff_pct}%', bg=C_LGREEN, bold=True, fc=C_GREEN if diff_pct < 5 else C_RED, align='center')
            remark = '單一來源' if len(valid) == 1 else ''
        elif len(valid) == 1:
            sup_name = list(valid.keys())[0]
            dc(ws, row, col,   sup_name, bg=bg, align='center', fc='555555')
            dc(ws, row, col+1, '—', bg=bg, align='center', fc='AAAAAA')
            remark = '單一來源'
        else:
            dc(ws, row, col,   '—', bg=bg, align='center', fc='AAAAAA')
            dc(ws, row, col+1, '—', bg=bg, align='center', fc='AAAAAA')
            remark = ''
        dc(ws, row, col+2, remark, bg=bg, fc='888888', sz=9, italic=True)
        ws.row_dimensions[row].height = 18
        row += 1

    # ── 欄寬 ──
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 6
    col = 5
    for i, sup in enumerate(suppliers):
        lc = get_column_letter(col)
        ws.column_dimensions[lc].width = 12     # 最新
        ws.column_dimensions[get_column_letter(col+1)].width = 10  # 最低
        ws.column_dimensions[get_column_letter(col+2)].width = 10  # 最高
        ws.column_dimensions[get_column_letter(col+3)].width = 8   # 次數
        col += 4
    ws.column_dimensions[get_column_letter(col)].width   = 14  # 最優廠商
    ws.column_dimensions[get_column_letter(col+1)].width = 10  # 價差%
    ws.column_dimensions[get_column_letter(col+2)].width = 12  # 備註
    ws.freeze_panes = 'A4'

# ────────────────────────────────────────────────
# 總覽比價表（所有品項，依大類分色）
# ────────────────────────────────────────────────
def write_overview(ws, df_all):
    CAT_INFO = {
        'BPP': ('外層布',  '1A5C3A'),
        'BEL': ('耳繩',   '5B2D8E'),
        'BMB': ('熔噴布',  '006D77'),
        'BES': ('親膚布',  'B45309'),
        'FBOX':('外箱',   'CC2200'),
        'BPA': ('包裝',   '1C3F6B'),
        'BAC': ('活性碳',  '36454F'),
        'BJF': ('接合布',  '36454F'),
        'BNW': ('鼻樑條',  '36454F'),
        'FPEB':('其他',   '555555'),
        'FKOP':('其他',   '555555'),
        'BPE': ('其他',   '555555'),
    }
    g = make_pivot(df_all)
    suppliers = sorted(g['廠商簡稱'].unique().tolist())
    n_sup = len(suppliers)
    total_cols = 5 + n_sup * 2 + 2

    # 標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws.cell(row=1, column=1,
        value=f'原物料進貨比價總覽  |  2024-01 ~ 2026-05  |  {g["品號"].nunique()} 品項  |  {n_sup} 家廠商')
    t.font = Font(name='微軟正黑體', bold=True, size=14, color='FFFFFF')
    t.fill = PatternFill('solid', start_color=C_DARK)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 廠商標題列
    col = 6
    for sup in suppliers:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        sc = ws.cell(row=2, column=col, value=sup)
        sc.font = Font(name='微軟正黑體', bold=True, size=10, color='FFFFFF')
        sc.fill = PatternFill('solid', start_color=C_DARK)
        sc.alignment = Alignment(horizontal='center', vertical='center')
        set_border(sc, 'medium', '888888')
        col += 2
    ws.row_dimensions[2].height = 20

    # 欄標題
    hdrs3 = ['大類','品號','品名','規格','單位']
    for ci, h in enumerate(hdrs3, 1):
        hd(ws, 3, ci, h, bg=C_DARK)
    col = 6
    for sup in suppliers:
        hd(ws, 3, col,   '最新進價', bg=C_DARK)
        hd(ws, 3, col+1, '最低進價', bg=C_DARK)
        col += 2
    hd(ws, 3, col,   '最優廠商', bg=C_GREEN)
    hd(ws, 3, col+1, '價差%',   bg=C_GREEN)
    ws.row_dimensions[3].height = 22

    # 依大類排序
    def get_cat(pno):
        for cat in CAT_INFO:
            if pno.startswith(cat):
                return cat
        return 'ZZZ'
    items = g.groupby(['品號','品名','規格','單位'])
    item_list = sorted(items.groups.keys(), key=lambda x: get_cat(x[0]) + x[0])

    row = 4
    prev_cat = ''
    for key in item_list:
        pno, pname, spec, unit = key
        cat = get_cat(pno)
        cat_label, cat_color = CAT_INFO.get(cat, ('其他','555555'))
        bg = C_LBLUE if row % 2 == 0 else C_WHITE

        # 大類換行
        if cat != prev_cat:
            prev_cat = cat
            bg_cat = cat_color
            # 插入小標題
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
            lc = ws.cell(row=row, column=1, value=f'▌ {cat_label}（{cat}）')
            lc.font = Font(name='微軟正黑體', bold=True, size=11, color='FFFFFF')
            lc.fill = PatternFill('solid', start_color=bg_cat)
            lc.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[row].height = 20
            row += 1
            bg = C_LBLUE if row % 2 == 0 else C_WHITE

        grp = items.get_group(key)
        dc(ws, row, 1, cat_label, bg=bg, fc=cat_color, bold=True, align='center', sz=9)
        dc(ws, row, 2, pno,       bg=bg, sz=9)
        dc(ws, row, 3, pname,     bg=bg, sz=10, bold=True)
        dc(ws, row, 4, spec,      bg=bg, sz=8, fc='666666')
        dc(ws, row, 5, unit,      bg=bg, align='center', sz=9)

        prices_by_sup = {}
        col = 6
        for sup in suppliers:
            r = grp[grp['廠商簡稱']==sup]
            if len(r) > 0:
                newest = r.iloc[0]['最新進價']
                lo     = r.iloc[0]['最低進價']
                prices_by_sup[sup] = newest
                dc(ws, row, col,   newest, bg=bg, align='right', fmt='#,##0.0000', sz=9)
                dc(ws, row, col+1, lo,     bg=bg, align='right', fmt='#,##0.0000', sz=9, fc=C_GREEN)
            else:
                dc(ws, row, col,   '—', bg=C_GRAY, align='center', fc='CCCCCC', sz=9)
                dc(ws, row, col+1, '—', bg=C_GRAY, align='center', fc='CCCCCC', sz=9)
            col += 2

        valid = {s:p for s,p in prices_by_sup.items() if pd.notna(p) and p > 0}
        if len(valid) >= 2:
            best_sup = min(valid, key=valid.get)
            worst_p  = max(valid.values())
            best_p   = valid[best_sup]
            diff_pct = round((worst_p - best_p) / worst_p * 100, 1)
            dc(ws, row, col,   best_sup, bg=C_LGREEN, bold=True, fc=C_GREEN, align='center', sz=9)
            dc(ws, row, col+1, f'{diff_pct}%', bg=C_LGREEN, bold=True,
               fc=C_RED if diff_pct > 10 else (C_ORANGE if diff_pct > 5 else C_GREEN), align='center', sz=9)
        else:
            dc(ws, row, col,   list(valid.keys())[0] if valid else '—', bg=bg, align='center', sz=9, fc='888888')
            dc(ws, row, col+1, '—', bg=bg, align='center', sz=9, fc='AAAAAA')

        ws.row_dimensions[row].height = 17
        row += 1

    # 欄寬
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 6
    col = 6
    for sup in suppliers:
        ws.column_dimensions[get_column_letter(col)].width = 11
        ws.column_dimensions[get_column_letter(col+1)].width = 10
        col += 2
    ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions[get_column_letter(col+1)].width = 9
    ws.freeze_panes = 'A4'

# ────────────────────────────────────────────────
# 建立所有工作表
# ────────────────────────────────────────────────

# 1. 總覽
ws_ov = wb.create_sheet('📊 總覽比價')
write_overview(ws_ov, df)

# 2. 各大類
cat_map = [
    ('🟢 BPP 外層布', 'BPP', '1A5C3A'),
    ('🟣 BEL 耳繩',   'BEL', '5B2D8E'),
    ('🔵 BMB 熔噴布', 'BMB', '006D77'),
    ('🟠 BES 親膚布', 'BES', 'B45309'),
    ('🔴 FBOX 外箱',  'FBOX','CC2200'),
    ('📦 BPA 包裝',   'BPA', '1C3F6B'),
    ('⬛ BAC 活性碳', 'BAC', '36454F'),
    ('⬛ BJF 接合布', 'BJF', '5A7A8A'),
    ('⬛ BNW 鼻樑條', 'BNW', '5A7A8A'),
    ('📎 其他材料',   None,  '555555'),
]

for sh_name, prefix, color in cat_map:
    ws = wb.create_sheet(sh_name)
    if prefix:
        sub = df[df['品號'].str.startswith(prefix)]
    else:
        sub = df[df['品號'].str.match(r'^(?!BPP|BEL|BMB|BES|FBOX|BPA|BAC|BJF|BNW)')]
    write_sheet(ws, sub, sh_name.replace('🟢','').replace('🟣','').replace('🔵','')
                .replace('🟠','').replace('🔴','').replace('📦','')
                .replace('⬛','').replace('📎','').strip(), color)

# 儲存
out = r'C:\Users\admin\OneDrive\桌面\原物料進貨比價表_20240102-20260508.xlsx'
wb.save(out)
print(f'OK: {out}')
print(f'工作表: {[s.title for s in wb.worksheets]}')
