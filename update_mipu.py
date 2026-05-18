import math, shutil, tempfile, os, sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
sys.stdout.reconfigure(encoding='utf-8')

SRC = r'C:\Users\admin\OneDrive\桌面\惠雯表格\秘笈-布料試算_NEW.xlsx'
# 備份 + 用 temp 路徑開啟（避免中文路徑鎖定問題）
bak = SRC.replace('.xlsx', '_bak.xlsx')
shutil.copy2(SRC, bak)
tmp = tempfile.mktemp(suffix='.xlsx')
shutil.copy2(SRC, tmp)

wb = load_workbook(tmp)
SH = 'TN95 3D、醫用3D Ultra外層布產出成品數量換算表'
ws = wb[SH]

# ── 樣式函數 ──
def bd(style='thin', color='AAAAAA'):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hd(ws, r, c, v, bg='1C3F6B', fc='FFFFFF', sz=11, bold=True, align='center', wrap=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name='微軟正黑體', bold=bold, size=sz, color=fc)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = bd('medium', '888888')
    return cell

def dc(ws, r, c, v, bg='FFFFFF', fc='222222', sz=10, bold=False, align='left', fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name='微軟正黑體', bold=bold, size=sz, color=fc)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = bd('hair', 'DDDDDD')
    if fmt: cell.number_format = fmt
    return cell

def merge_row(ws, r, c1, c2, v, bg, fc='FFFFFF', sz=11, bold=True, align='left'):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=v)
    cell.font = Font(name='微軟正黑體', bold=bold, size=sz, color=fc)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = bd('medium', '888888')
    ws.row_dimensions[r].height = 22
    return cell

# ── 從第 25 列開始新增內容（保留上面原有資料） ──
R = 25

# 先清除 R=25 以後的合併儲存格與內容，避免 MergedCell 衝突
merged_to_remove = [rng for rng in ws.merged_cells.ranges
                    if rng.min_row >= R]
for rng in merged_to_remove:
    ws.unmerge_cells(str(rng))
# 清除儲存格內容
for row in ws.iter_rows(min_row=R, max_row=ws.max_row):
    for cell in row:
        cell.value = None

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 分隔線
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
merge_row(ws, R, 1, 6,
    '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━',
    '555555', sz=9)
R += 1

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 大標題
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws.row_dimensions[R].height = 30
merge_row(ws, R, 1, 6,
    '★  DRX-3D 兒童印刷外層布  叫料換算秘笈  ★  （老闆方法 L=13.5cm / M=12.5cm 實測確認 | XL.S 待確認 | 損耗0.4%）',
    '5B2D8E', sz=13, align='center')
R += 1

# 警告列：舊算法 vs 新算法
ws.row_dimensions[R].height = 20
merge_row(ws, R, 1, 6,
    '⚠  舊算法：M=730片/100m（步距13.7cm，損耗等同抓13%，過多！）  →  正確：M=800片/100m(12.5cm) ／ L=741片/100m(13.5cm)',
    'FFF0F0', fc='CC2200', sz=10)
R += 1

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 換算基準表
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
hdrs1 = ['成品規格', 'Pitch步距', '每100M片數', '1000M捲\n(0.4%損耗)', '800M捲\n(0.4%損耗)', '備註']
for ci, h in enumerate(hdrs1, 1):
    hd(ws, R, ci, h, bg='5B2D8E')
ws.row_dimensions[R].height = 28
R += 1

# 各尺寸 Pitch 對照（L=13.5cm已確認，M=12.5cm已確認，XL/S待確認）
def calc_roll(pitch, roll_m, loss=0.004):
    pieces = int(roll_m * 100 / pitch * (1 - loss))
    boxes  = pieces // 20
    return f'{pieces:,}片 ≈ {boxes}盒'

size_rows = [
    ('XL 尺寸', '待確認', '—', '—', '—', '⏳ 待HANK量測'),
    ('L 尺寸（兒童）', '13.5 cm', f'{100*100//135*1 if False else round(100*100/13.5)}片',
     calc_roll(13.5, 1000), calc_roll(13.5, 800), '✅ 2026-05-18確認'),
    ('M 尺寸（兒童）', '12.5 cm', '800片',
     calc_roll(12.5, 1000), calc_roll(12.5, 800), '✅ 2026-05-18確認'),
    ('S 尺寸', '待確認', '—', '—', '—', '⏳ 待HANK量測'),
]
# 重新計算 L 每100M片數
size_rows[1] = ('L 尺寸（兒童）', '13.5 cm',
                f'{int(100*100/13.5)}片',
                calc_roll(13.5, 1000), calc_roll(13.5, 800), '✅ 2026-05-18確認')

row_bgs = [
    (['FAFAFA','DDDDDD','DDDDDD','DDDDDD','DDDDDD','DDDDDD'],
     ['888888','888888','888888','888888','888888','888888']),  # XL 待確認灰色
    (['F3EEF8','F3EEF8','FFF9FF','D4EFD4','D4EFD4','E8F5E9'],
     ['222222','1E8449','1E8449','1A5C3A','1A5C3A','1A5C3A']),  # L 綠色
    (['F3EEF8','F3EEF8','FFF9FF','E8D5F5','E8D5F5','D6E8FB'],
     ['222222','5B2D8E','5B2D8E','1C3F6B','1C3F6B','1C3F6B']),  # M 藍紫色
    (['FAFAFA','DDDDDD','DDDDDD','DDDDDD','DDDDDD','DDDDDD'],
     ['888888','888888','888888','888888','888888','888888']),  # S 待確認灰色
]
for (vals, (bgs, fcs)) in zip(size_rows, row_bgs):
    for ci, (v, bg, fc) in enumerate(zip(vals, bgs, fcs), 1):
        dc(ws, R, ci, v, bg=bg, fc=fc, bold=True, align='center', sz=10)
    ws.row_dimensions[R].height = 22
    R += 1
R += 1

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 計算步驟
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws.row_dimensions[R].height = 22
merge_row(ws, R, 1, 6, '📌  正確計算步驟（老闆教的，不要分開算！）', '1C3F6B', align='left')
R += 1

steps = [
    ('STEP 1', '同品相跨通路訂單盒數 全部加總（不要按通路分開算，會有誤差！）',
     'D6E8FB', '1C3F6B'),
    ('STEP 2', '總盒數 × 20片/盒 = 總需求片數',
     'EEF5FF', '222222'),
    ('STEP 3', '需求M數 = 總片數 × Pitch(cm) ÷ 100 ÷ 0.996（除掉0.4%損耗）｜M用12.5 / L用13.5 / XL.S待確認',
     'D6E8FB', '222222'),
    ('STEP 4', '捲數 = 需求M數 ÷ 捲長(1000M或800M)，無條件進位（不夠就整捲補）',
     'EEF5FF', '222222'),
    ('STEP 5', '找片數最多的品相定捲數，其他品相依需求取1,200M整數倍叫料',
     'D6E8FB', '222222'),
]
for step_label, step_text, step_bg, step_fc in steps:
    dc(ws, R, 1, step_label, bg='1C3F6B', fc='FFFFFF', bold=True, align='center', sz=10)
    ws.merge_cells(start_row=R, start_column=2, end_row=R, end_column=6)
    cell = ws.cell(row=R, column=2, value=step_text)
    cell.font = Font(name='微軟正黑體', size=10, color=step_fc)
    cell.fill = PatternFill('solid', start_color=step_bg)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = bd('hair', 'DDDDDD')
    ws.row_dimensions[R].height = 20
    R += 1

R += 1

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 快速查表
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws.row_dimensions[R].height = 22
merge_row(ws, R, 1, 6,
    '📊  快速查表（M尺寸 Pitch=12.5cm ／ L尺寸 Pitch=13.5cm ／ 捲長1000M ／ 損耗0.4%）',
    '1A5C3A', align='center')
R += 1

ref_hdrs = ['訂單盒數\n（加總）', '需求片數',
            'M號需M數\n(12.5cm/0.4%)', 'M號需捲數\n(1000M)',
            'L號需M數\n(13.5cm/0.4%)', 'L號需捲數\n(1000M)']
for ci, h in enumerate(ref_hdrs, 1):
    hd(ws, R, ci, h, bg='1A5C3A')
ws.row_dimensions[R].height = 30
R += 1

PITCH_M  = 12.5
PITCH_L  = 13.5
LOSS     = 0.996   # 0.4%損耗
ROLL_M   = 1000
PER_BOX  = 20

box_targets = [100, 200, 300, 485, 600, 800, 1000, 1200, 1464, 1584, 1824, 2000, 2500, 3000]
ref_a = 'E8F5E9'; ref_b = 'FFFFFF'
for i, boxes in enumerate(box_targets):
    pieces_need = boxes * PER_BOX

    m_need_M  = pieces_need * PITCH_M / 100 / LOSS
    rolls_M   = math.ceil(m_need_M / ROLL_M)

    m_need_L  = pieces_need * PITCH_L / 100 / LOSS
    rolls_L   = math.ceil(m_need_L / ROLL_M)

    bg = ref_a if i % 2 == 0 else ref_b
    highlight = boxes in (1464, 1584, 1824)
    if highlight:
        bg = 'FFFDE7'

    dc(ws, R, 1, boxes,               bg=bg, align='right', bold=highlight, fc='1C3F6B' if highlight else '222222', sz=10)
    dc(ws, R, 2, pieces_need,         bg=bg, align='right', fmt='#,##0')
    dc(ws, R, 3, math.ceil(m_need_M), bg=bg, align='right', fmt='#,##0', fc='1F6FBF')
    dc(ws, R, 4, rolls_M,             bg=bg, align='center', bold=True, fc='5B2D8E', sz=12)
    dc(ws, R, 5, math.ceil(m_need_L), bg=bg, align='right', fmt='#,##0', fc='1E8449')
    dc(ws, R, 6, rolls_L,             bg=bg, align='center', bold=True, fc='1A5C3A', sz=12)
    ws.row_dimensions[R].height = 18
    R += 1

R += 1

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 本次訂單試算（5/15叫料）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws.row_dimensions[R].height = 22
merge_row(ws, R, 1, 6,
    '🚀  本次訂單試算（5/15叫料，M尺寸，20片/盒）',
    'CC7700', align='left')
R += 1

order_hdrs = ['品項', '啄木鳥', '家樂福', '躍獅', '合計盒數', '合計片數']
for ci, h in enumerate(order_hdrs, 1):
    hd(ws, R, ci, h, bg='CC7700')
ws.row_dimensions[R].height = 20
R += 1

orders = [
    ('小車與飛行隊', 360, 1464, 0),
    ('卡皮巴拉',     0,   1464, 120),
    ('粉嫩獨角獸',   0,   1464, 0),
    ('企鵝寶寶',     120, 0,    0),
    ('花花兔',       120, 0,    0),
]
order_bg_a = 'FFF8E1'; order_bg_b = 'FFFDE7'
for i, (name, ww, klf, ys) in enumerate(orders):
    total_boxes  = ww + klf + ys
    total_pieces = total_boxes * PER_BOX
    bg = order_bg_a if i % 2 == 0 else order_bg_b
    max_item = (name == '小車與飛行隊')

    dc(ws, R, 1, name,         bg=bg, bold=max_item, fc='1C3F6B' if max_item else '222222')
    dc(ws, R, 2, ww  if ww  else '—', bg=bg, align='center', fc='555555' if not ww else '222222')
    dc(ws, R, 3, klf if klf else '—', bg=bg, align='center', fc='555555' if not klf else '222222')
    dc(ws, R, 4, ys  if ys  else '—', bg=bg, align='center', fc='555555' if not ys else '222222')
    dc(ws, R, 5, total_boxes,  bg=bg, align='right', bold=True, fc='CC7700', fmt='#,##0')
    dc(ws, R, 6, total_pieces, bg=bg, align='right', bold=max_item, fc='5B2D8E', fmt='#,##0')
    ws.row_dimensions[R].height = 18
    R += 1

# 合計行
total_all_pieces = sum((w+k+y)*20 for _, w, k, y in orders)
dc(ws, R, 1, '合計',           bg='CC7700', fc='FFFFFF', bold=True, align='center')
dc(ws, R, 2, '',               bg='CC7700')
dc(ws, R, 3, '',               bg='CC7700')
dc(ws, R, 4, '',               bg='CC7700')
dc(ws, R, 5, '',               bg='CC7700')
dc(ws, R, 6, f'共 {total_all_pieces:,} 片', bg='CC7700', fc='FFFFFF', bold=True, align='right', sz=11)
ws.row_dimensions[R].height = 20
R += 2

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 本次叫料計算結果
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ws.row_dimensions[R].height = 22
merge_row(ws, R, 1, 6, '✅  本次叫料結果（老闆算法）', '1C3F6B', align='left')
R += 1

result_hdrs = ['品項', '合計片數', '需求M數(+3%)', '需叫捲數', '庫存', '實際應叫']
for ci, h in enumerate(result_hdrs, 1):
    hd(ws, R, ci, h, bg='1C3F6B')
ws.row_dimensions[R].height = 20
R += 1

order_results = [
    ('小車與飛行隊', (360+1464)*20, '無庫存'),
    ('卡皮巴拉',     (1464+120)*20, '快5捲→約2,425盒 庫存足'),
    ('粉嫩獨角獸',   1464*20,       '無庫存'),
    ('企鵝寶寶',     120*20,        '無庫存'),
    ('花花兔',       120*20,        '無庫存'),
]
res_bg_a = 'D6E8FB'; res_bg_b = 'EEF5FF'
for i, (name, pieces, stock_note) in enumerate(order_results):
    m_need = pieces * PITCH_M / 100 / LOSS
    rolls  = math.ceil(m_need / ROLL_M)
    bg = res_bg_a if i % 2 == 0 else res_bg_b
    has_stock = '庫存足' in stock_note

    dc(ws, R, 1, name,              bg=bg, bold=True, fc='1C3F6B')
    dc(ws, R, 2, pieces,            bg=bg, align='right', fmt='#,##0', fc='5B2D8E')
    dc(ws, R, 3, int(m_need+1),     bg=bg, align='right', fmt='#,##0', fc='CC7700')
    dc(ws, R, 4, rolls,             bg=bg, align='center', bold=True, fc='5B2D8E', sz=13)
    dc(ws, R, 5, stock_note,        bg=bg, align='left',  fc='1A5C3A' if has_stock else 'CC2200', sz=9)
    dc(ws, R, 6, '0捲（庫存覆蓋）' if has_stock else f'{rolls}捲',
       bg='E8F5E9' if has_stock else bg,
       align='center', bold=True,
       fc='1A5C3A' if has_stock else 'CC2200', sz=11)
    ws.row_dimensions[R].height = 22
    R += 1

# 最後備註
R += 1
ws.row_dimensions[R].height = 20
merge_row(ws, R, 1, 6,
    '💡 老闆提醒：片數最多的品相（小車與飛行隊）先計算捲數，其他品相以1,200M整倍數對齊 | 卡皮巴拉庫存約5捲(≈2,425盒)，本次不需叫布',
    'F5F5F5', fc='444444', sz=9)

# ── 欄寬調整 ──
for col, w in zip('ABCDEF', [20, 12, 14, 14, 14, 22]):
    ws.column_dimensions[col].width = w

OUT = r'C:\Users\admin\OneDrive\桌面\惠雯表格\秘笈-布料試算_NEW.xlsx'
wb.save(tmp)
shutil.copy2(tmp, OUT)
os.unlink(tmp)
print(f'OK: {OUT}')
print('備份原檔:', bak)
