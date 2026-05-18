# -*- coding: utf-8 -*-
"""
缺料分析自動化腳本
每日執行: python 缺料分析.py
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import re, os, glob, sys, io, math
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ─── 設定 ─────────────────────────────────────────────────
BASE   = r"C:\Users\admin\OneDrive\桌面\CODE資料"
DESK   = r"C:\Users\admin\OneDrive\桌面"
today  = datetime.now().strftime("%m%d")
outdir = os.path.join(BASE, f"缺料分析{today}")
os.makedirs(outdir, exist_ok=True)

def unique_path(base_path):
    """若檔案已存在（可能被鎖住），自動加流水號"""
    if not os.path.exists(base_path):
        return base_path
    root, ext = os.path.splitext(base_path)
    for i in range(1, 100):
        p = f"{root}_{i}{ext}"
        if not os.path.exists(p):
            return p
    return base_path

def latest(pat):
    f = glob.glob(pat)
    return max(f, key=os.path.getmtime) if f else None

def latest2(*pats):
    """從多個 glob pattern 中找最新的檔案"""
    files = []
    for pat in pats:
        files.extend(glob.glob(pat))
    return max(files, key=os.path.getmtime) if files else None

inv_path    = latest2(os.path.join(DESK, "原物料庫存*.xlsx"),
                      os.path.join(BASE, "原物料庫存*.xlsx"))
ear_path    = latest2(os.path.join(DESK, "耳繩庫存*.xls"),
                      os.path.join(DESK, "耳繩庫存*.xlsx"),
                      os.path.join(BASE, "耳繩庫存*.xls"),
                      os.path.join(BASE, "耳繩庫存*.xlsx"))
box_path    = latest2(os.path.join(DESK, "盒子庫存*.xls"),
                      os.path.join(BASE, "盒子庫存*.xls"))
bom_path    = latest(os.path.join(BASE, "製令總表_合併*.xlsx"))
_CLD = r"C:\Users\admin\OneDrive\桌面\Claude紀錄"
order_path  = (latest(os.path.join(DESK, "訂單生產總表 *.xlsx"))
            or latest(os.path.join(DESK, "訂單生產總表*.xlsx"))
            or latest(os.path.join(BASE, "訂單生產總表*.xlsx"))
            or latest(os.path.join(_CLD, "訂單生產總表 *.xlsx"))
            or latest(os.path.join(_CLD, "訂單生產總表*.xlsx")))
# 主BOM：優先用 BOM20260323（最完整），其次原物料用途清單
CLAUDE_DIR  = r"C:\Users\admin\OneDrive\桌面\Claude紀錄"
master_path = (latest(os.path.join(CLAUDE_DIR, "BOM*.xlsx"))
            or latest(os.path.join(DESK, "*原物料用途清單*.xlsx"))
            or latest(os.path.join(BASE, "*原物料用途清單*.xlsx")))
# INVI02：品號→品名 對照表（ERP系統品名）
invi_path   = (latest(os.path.join(DESK, "INVI02*.XLSX"))
            or latest(os.path.join(DESK, "INVI02*.xlsx"))
            or latest(os.path.join(BASE, "INVI02*.XLSX"))
            or latest(os.path.join(BASE, "INVI02*.xlsx")))
# 原料類品號：品號→品名（比 INVI02 更完整的原料名稱對照）
rawmat_path = (latest(os.path.join(DESK, "原料類品號*.XLSX"))
            or latest(os.path.join(DESK, "原料類品號*.xlsx"))
            or latest(os.path.join(BASE, "原料類品號*.XLSX"))
            or latest(os.path.join(BASE, "原料類品號*.xlsx")))

print("=== 找到檔案 ===")
for lbl, p in [("原物料庫存", inv_path), ("耳繩庫存",  ear_path),
               ("盒子庫存",   box_path), ("BOM製令",   bom_path),
               ("訂單",       order_path),
               ("主BOM",      master_path),
               ("INVI02品名", invi_path),
               ("原料類品號", rawmat_path)]:
    print(f"  [{lbl}] {os.path.basename(p) if p else '❌ 未找到'}")

if not all([inv_path, bom_path, order_path]):
    print("\n❌ 缺少必要檔案，請確認後重試"); sys.exit(1)

# ─── 品號→品名 對照表（優先用原料類品號，其次 INVI02）────────
invi_name = {}

def _load_name_table(path, header_row, cols):
    try:
        df = pd.read_excel(path, header=header_row, usecols=cols)
        return {str(r[cols[0]]).strip(): str(r[cols[1]]).strip()
                for _, r in df.iterrows()
                if pd.notna(r[cols[0]]) and pd.notna(r[cols[1]])}
    except Exception as e:
        print(f"  ⚠ 讀取失敗 {os.path.basename(path)}: {e}")
        return {}

if rawmat_path:
    invi_name = _load_name_table(rawmat_path, 2, ['品號', '品名'])
    print(f"原料類品號對照: {len(invi_name)} 筆")
elif invi_path:
    invi_name = _load_name_table(invi_path, 2, ['品號', '品名'])
    print(f"INVI02 品名對照: {len(invi_name)} 筆")
else:
    print("  ⚠ 未找到品名對照檔案（INVI02 / 原料類品號）")

# ─── 材料品號分類（需在計算前定義）───────────────────────
def mat_code_type(code):
    code = str(code)
    if code.startswith('BEL'):  return '耳繩'
    if code.startswith('BMB'):  return '熔噴'
    if code.startswith('BES'):  return '親膚'
    # BPP-FM-ZA* / BPP-FM-XA* = 印花外層（有圖案設計，隨訂單叫貨，不在庫存）
    if code.startswith('BPP-FM-ZA') or code.startswith('BPP-FM-XA'): return '印花外層'
    if code.startswith('BPP'):  return '外層'
    # BPE-FM-* / BPET-* = 印花外層不織布（隨訂單叫貨，不在庫存檔）
    if code.startswith('BPE') or code.startswith('BPET'): return '印花外層'
    # BJF = 彈性布（3D彈力口罩用，隨訂單叫貨）
    if code.startswith('BJF'): return '印花外層'
    # BPA = 包裝盒（盒子）
    if code.startswith('BPA'): return '盒子'
    return '其他'

# ─── 1. 解析原物料庫存 ────────────────────────────────────
def parse_qty(qty_str):
    """從 '28 34K 14M' 或 '5B 10K 3M' 取出總量（避免重複計算）
    與 parse_fabric_qty 邏輯相同：先把倉別代碼移除，剩餘才是主倉數字。
    '12M'  → 12  ('12' 直接後接 M，整體歸入倉別，主倉=0)
    '5 12M'→ 17  (5 主倉 + 12 M倉)
    """
    s = str(qty_str).strip()
    # 有倉別字母的數量（K M B等）
    extra = sum(float(x) for x in re.findall(r'(\d+(?:\.\d+)?)[KkMmBb]', s))
    # 主倉數字：先移除所有倉別代碼，剩餘的才是主倉（避免 re.match 回溯問題）
    s_no_extra = re.sub(r'\d+(?:\.\d+)?[KkMmBb]', '', s)
    m = re.search(r'(\d+(?:\.\d+)?)', s_no_extra)
    main = float(m.group(1)) if m else 0
    return main + extra

def parse_fabric_qty(qty_str):
    """外層布/熔噴布專用：解析完整捲數（含各倉別代號）
    K/M/B 均為倉別代號（完整捲數），不是米/千米單位。
    主數字後緊接倉別字母 → 該數字歸入倉別，不重複計算。

    '5  42M'      → 5+42  = 47  捲  (主倉5 + M倉42)
    '4  2K 46M'   → 4+2+46= 52  捲  (主倉4 + K倉2 + M倉46)
    '14  15M'     → 14+15 = 29  捲  (主倉14 + M倉15)
    '12M'         → 0+12  = 12  捲  (M倉12，無獨立主倉)
    '14M'         → 0+14  = 14  捲  (M倉14，無獨立主倉)
    '22M'         → 0+22  = 22  捲  (M倉22，無獨立主倉)
    '37'          → 37         捲
    '0  18M'      → 0+18  = 18  捲  (M倉18)

    ⚠ 實作注意：不可用 re.match(r'\\d+(?![KMB])') 的 negative lookahead，
    因為 regex 引擎會回溯：'14M' 先嘗試 '14' 失敗，再退回 '1' 成功 → 錯誤地算 1+14=15。
    正確做法：先 re.sub 移除所有 '數字+倉別字母'，剩餘字串再找第一個數字當主倉。
    """
    # 去掉括號備注（如 (190可替代)、(先用完)、(3D彈力淡紫用) 等）
    s = re.sub(r'\([^)]*\)', '', str(qty_str)).strip()
    # 倉別捲數（K/M/B 後綴數字，全部加總）
    warehouse_vals = re.findall(r'(\d+(?:\.\d+)?)\s*[KkMmBb]', s)
    warehouse_rolls = sum(float(x) for x in warehouse_vals)
    # 主倉數字：先把所有倉別代碼（數字+K/M/B）從字串移除，
    # 剩餘的第一個數字才是主倉捲數（避免 re.match 回溯把 '14M' 匹配成 1+14=15）
    s_no_wh = re.sub(r'\d+(?:\.\d+)?\s*[KkMmBb]', '', s)
    m0 = re.search(r'(\d+(?:\.\d+)?)', s_no_wh)
    main_rolls = float(m0.group(1)) if m0 else 0.0
    return main_rolls + warehouse_rolls, 0.0   # 無剩餘米數追蹤

import shutil, tempfile
_inv_tmp = os.path.join(tempfile.gettempdir(), 'inv_tmp_' + os.path.basename(inv_path))
try:
    shutil.copy2(inv_path, _inv_tmp)
    _inv_load_path = _inv_tmp
except Exception:
    _inv_load_path = inv_path
wb_inv = openpyxl.load_workbook(_inv_load_path, data_only=True)
ws_inv = wb_inv.active

SKIP_VALS = {'熔噴', '外層', '親膚', '耳線', '需叫貨', ''}
inv_table = []   # [{'gsm','width','kw','qty','raw','cat'}]

# ─── 解析盒子庫存（A/B欄：名稱+數量；D欄：名稱*數量）────────────
box_inventory = {}   # 盒子中文名 → 總數量
if box_path:
    try:
        import xlrd as _xlrd
        _wb_box = _xlrd.open_workbook(box_path)
        _ws_box = _wb_box.sheet_by_index(0)
        for _br in range(_ws_box.nrows):
            _row_b = [str(_ws_box.cell_value(_br, _bc) if _ws_box.ncols > _bc else '').strip()
                      for _bc in range(max(4, _ws_box.ncols))]
            # 左欄 A+B：盒子名稱 + 數量（小數點去掉 .0）
            _bname, _bqty_s = _row_b[0], _row_b[1]
            if _bname and _bname not in ('盒子庫存',) and _bqty_s:
                try:
                    box_inventory[_bname] = box_inventory.get(_bname, 0) + float(_bqty_s)
                except ValueError:
                    pass
            # 右欄 D：「名稱*數量」格式
            _dcell = _row_b[3] if len(_row_b) > 3 else ''
            _dm = re.match(r'^(.+?)\s*\*\s*(\d+(?:\.\d+)?)$', _dcell)
            if _dm:
                _bn2, _bq2 = _dm.group(1).strip(), float(_dm.group(2))
                box_inventory[_bn2] = box_inventory.get(_bn2, 0) + _bq2
        print(f'盒子庫存解析: {len(box_inventory)} 種')
        for _bk, _bv in sorted(box_inventory.items()):
            print(f'  {_bk}: {_bv:.0f}')
    except Exception as _e:
        print(f'  ⚠ 盒子庫存讀取失敗: {_e}')
else:
    print('  ⚠ 未找到盒子庫存檔案')

# ── 預先掃描 Row1 識別「需叫貨」欄（不計入現有庫存）
_all_rows = list(ws_inv.iter_rows(values_only=True))
_header_row = _all_rows[0] if _all_rows else []
SKIP_COLS = set()  # 需跳過的欄位 index（0-based）
for _ci, _hv in enumerate(_header_row):
    if isinstance(_hv, str) and '需叫貨' in _hv:
        SKIP_COLS.add(_ci)

# ── 預先掃描 B 欄取出耳繩段落資料（3737/737/73714/1610）
_EAR_SECTIONS = {737, 3737, 73714, 1610}
_ear_pre = []    # 暫存，之後合入 ear_table
_ear_sec = None
# ⚠️ 喬煒耳繩庫存單位換算（2026-05-05）：
#    原物料庫存0504.xlsx 耳繩段落（3737/737/73714/1610）記錄為【箱數】
#    實測：6箱 ≈ 16.27~17.05 kg，取平均值 16.66kg/6箱 = 2.777 kg/箱
_EAR_KG_PER_BOX = (16.27 + 17.05) / 2 / 6   # ≈ 2.777 kg/箱
for _row in ws_inv.iter_rows(values_only=True):
    if len(_row) < 2:
        continue
    _bv = _row[1]   # B 欄（index=1）
    if isinstance(_bv, int) and _bv in _EAR_SECTIONS:
        _ear_sec = str(_bv)
    elif isinstance(_bv, str) and _ear_sec:
        _v = _bv.strip()
        if re.match(r'^\d+\*\d+', _v):   # 遇到布料格式則離開耳繩段落
            _ear_sec = None
        else:
            _me = re.match(r'^(\S+)\s*\*\s*(\d+(?:\.\d+)?)', _v)
            if _me and not re.fullmatch(r'[A-Za-z]', _me.group(1)):
                _boxes = float(_me.group(2))
                _qty_kg = round(_boxes * _EAR_KG_PER_BOX, 2)
                _raw_label = f'{_me.group(1)}*{_boxes:.0f}箱({_qty_kg:.1f}kg)'
                _ear_pre.append({'code': _me.group(1), 'qty': _qty_kg,
                                 'raw': _raw_label, 'section': _ear_sec})

for row in _all_rows:
    for _ci, val in enumerate(row):
        if _ci in SKIP_COLS:   # 跳過「需叫貨」欄（不計入現有庫存）
            continue
        if not isinstance(val, str):
            continue
        v = val.strip()
        if v in SKIP_VALS:
            continue

        # ── 親膚格式: gsm*width*qty[B]?(brand) [XXK嘉]
        # e.g. "20*175*44(福B) 48K嘉", "20*220*3B(嘉) 16K"
        m_skin = re.match(r'^(\d+)\*(\d+)\*(\d+(?:\.\d+)?)[Bb]?\(([^)]*)\)(.*)', v)
        if m_skin:
            gsm       = int(m_skin.group(1))
            width     = int(m_skin.group(2))
            qty_b     = float(m_skin.group(3))
            brand_raw = m_skin.group(4)
            rest      = m_skin.group(5)

            if '福' in brand_raw:
                inv_table.append({'gsm': gsm, 'width': width, 'kw': '福綿',
                                  'qty': qty_b, 'raw': v, 'cat': '親膚'})
            # 嘉谷量：rest 中所有倉別數量(1B+17M+...)加總，再補 digit+嘉 純數字(如8嘉M)
            # 範例: "1B 17M嘉" → 1+17=18; "48M嘉" → 48; "8嘉M" → 8
            jg_wh = sum(float(x) for x in re.findall(r'(\d+(?:\.\d+)?)[KkMmBb]', rest))
            rest_no_wh = re.sub(r'\d+(?:\.\d+)?[KkMmBb]', '', rest)
            jg_plain = sum(float(x) for x in re.findall(r'(\d+(?:\.\d+)?)嘉', rest_no_wh))
            jg_qty = jg_wh + jg_plain
            if jg_qty > 0:
                inv_table.append({'gsm': gsm, 'width': width, 'kw': '嘉谷',
                                  'qty': jg_qty,
                                  'raw': v, 'cat': '親膚'})
            elif '嘉' in brand_raw:
                # 嘉谷專屬：括號內數量 + rest 中其他倉別數量（如 16K）
                rest_extra = sum(float(x) for x in re.findall(r'(\d+(?:\.\d+)?)[KkMmBb]', rest))
                inv_table.append({'gsm': gsm, 'width': width, 'kw': '嘉谷',
                                  'qty': qty_b + rest_extra, 'raw': v, 'cat': '親膚'})
            continue

        # ── 標準格式: gsm*width[keyword]*qty [XK XM ...]
        # e.g. "30*175藍*28 34K 14M", "20*175醫灰*42"
        m_std = re.match(r'^(\d+)\*(\d+)([^*]*)\*(.+)$', v)
        if not m_std:
            continue
        gsm     = int(m_std.group(1))
        width   = int(m_std.group(2))
        keyword = m_std.group(3).strip()
        # 外層/熔噴：分離完整捲數與剩餘米數
        qty_rolls, qty_m = parse_fabric_qty(m_std.group(4))
        qty = qty_rolls   # 向後相容（find_stock 內部改用 qty_rolls + qty_m/rl）

        # 分類
        if 'H12' in keyword:
            cat, kw = '熔噴', 'H12'
        elif 'H11' in keyword:
            cat, kw = '熔噴', 'H11'
        elif '醫灰' in keyword or ('灰' in keyword and gsm <= 25):
            cat, kw = '熔噴', '醫灰'
        elif '醫' in keyword and '白' in keyword:   # 醫白/醫用白/白醫 → 明確醫用白
            cat, kw = '熔噴', '醫白'
        elif keyword == '' and gsm <= 25:
            cat, kw = '熔噴', '醫白'
        elif gsm >= 28:
            cat = '外層'
            kw  = keyword if keyword else '無色'
        else:
            continue   # 無法分類跳過

        inv_table.append({'gsm': gsm, 'width': width, 'kw': kw,
                          'qty': qty, 'qty_rolls': qty_rolls, 'qty_m': qty_m,
                          'raw': v, 'cat': cat})

print(f"原物料庫存解析: {len(inv_table)} 筆")
for it in inv_table:
    qr = it.get('qty_rolls', it['qty'])
    qm = it.get('qty_m', 0)
    extra = f" +{qm:.0f}M" if qm > 0 else ""
    print(f"  {it['cat']:3} gsm={it['gsm']:2} w={it['width']:3} kw={it['kw']:6} qty={qr:7.1f}捲{extra:8}  [{it['raw']}]")

# ─── 耳繩庫存 ─────────────────────────────────────────────
ear_table = []
ear_table.extend(_ear_pre)   # 合入原物料庫存 B 欄的耳繩段落資料
def _parse_luyuan_qty(qty_str):
    """解析綠沅庫存欄：'B*3.5  M*6' → 9.5；'庫2' → 2；'庫0' → 0"""
    s = str(qty_str).strip()
    parts = re.findall(r'[A-Za-z]\s*\*\s*(\d+(?:\.\d+)?)', s)
    if parts: return sum(float(x) for x in parts)
    m = re.search(r'庫\s*(\d+(?:\.\d+)?)', s)
    if m: return float(m.group(1))
    m = re.match(r'(\d+(?:\.\d+)?)', s)
    return float(m.group(1)) if m else 0.0

if ear_path:
    try:
        engine = 'xlrd' if str(ear_path).endswith('.xls') else None
        xl_e = pd.ExcelFile(ear_path, engine=engine)
        for sn in xl_e.sheet_names:
            df_e = xl_e.parse(sn, header=None)

            # ── 綠沅工作表：格式 [綠沅6.0-色名碼] [B*X M*Y / 庫N] [備註]
            if '綠沅' in sn:
                for _, row in df_e.iterrows():
                    row_list = ['' if pd.isna(v) else v for v in row]
                    name_cell = str(row_list[0]).strip()
                    qty_cell  = str(row_list[1]).strip() if len(row_list) > 1 else ''
                    if not name_cell or '耳繩庫存' in name_cell or name_cell in ('品名', ''):
                        continue
                    nm = re.match(r'綠沅\d+\.?\d*[-－](.+)', name_cell)
                    if not nm:
                        continue
                    color_code = nm.group(1).strip()   # e.g. 淡紫235, 黑元筋
                    qty = _parse_luyuan_qty(qty_cell)
                    ear_table.append({'code': color_code,
                                      'qty':  qty,
                                      'raw':  name_cell,
                                      'section': '綠沅6.0'})
                    print(f"  綠沅  code={color_code}  qty={qty}  [{name_cell}]")
                continue   # 綠沅已處理，跳過下方通用邏輯

            # ── 喬煒等一般工作表 ──
            cur_section = None
            for _, row in df_e.iterrows():
                row_list = list(row)
                for ci, cell in enumerate(row_list):
                    if isinstance(cell, (int, float)) and not pd.isna(cell):
                        v = float(cell)
                        if v == int(v) and int(v) > 100:
                            cur_section = str(int(v))
                        continue
                    if not isinstance(cell, str):
                        continue
                    ce = cell.strip()
                    if not ce:
                        continue
                    # 格式1: "412珊瑚橘*3" 排除單一大寫字母（倉別 B*3 M*7）
                    m = re.match(r'^(\S+)\s*\*\s*(\d+(?:\.\d+)?)', ce)
                    if m and m.group(1).strip() and not re.fullmatch(r'[A-Za-z]', m.group(1).strip()):
                        ear_table.append({'code': m.group(1).strip(),
                                          'qty':  float(m.group(2)),
                                          'raw':  ce,
                                          'section': cur_section or ''})
                        continue
                    # 格式2: "335珊瑚橘" + 下欄數字
                    if re.match(r'^\d{3}\w', ce):
                        if ci + 1 < len(row_list):
                            nxt = row_list[ci + 1]
                            if isinstance(nxt, (int, float)) and not pd.isna(nxt) and nxt > 0:
                                ear_table.append({'code': ce,
                                                  'qty':  float(nxt),
                                                  'raw':  f'{ce}*{nxt}',
                                                  'section': cur_section or ''})
    except Exception as e:
        print(f"  耳繩庫存讀取錯誤: {e}")

print(f"耳繩庫存解析:   {len(ear_table)} 筆")
for it in ear_table:
    print(f"  code={it['code']}  qty={it['qty']}  [{it['raw']}]")

# ─── 2. 庫存查找 ─────────────────────────────────────────
# 複合色判斷：若 kw 含有 color_kw 以外的其他顏色字，視為複合色不做模糊匹配
_COLOR_CHARS = set('灰藍紫粉橘綠黃白黑棕紅')

def _is_compound_color(kw, ckw):
    """kw 中含有 ckw 以外的顏色字 → 複合色，不做子字串匹配"""
    other = _COLOR_CHARS - set(ckw)
    return any(c in kw for c in other)

# 寬度替代規則：當庫存中無此寬度時，可嘗試的替代寬度
WIDTH_SUBS = {
    195: 200,   # 195mm 外層可用 200mm 替代（如福綿195抹茶→40*200抹茶）
    250: 260,   # 250mm 親膚可用 260mm 替代（如福綿250白→20*260福B）
}

# 特定品號指定替代庫存（BOM品號 → 實際查找參數）
# 用途：BOM品號規格與庫存格式不符時，手動指定對應庫存
BOM_OVERRIDES = {
    # BPP-JD220-PUR001-040（薰衣草紫220mm）→ 庫存用 40gsm*250mm*薰衣草 替代
    'BPP-JD220-PUR001-040': {'cat': '外層',  'gsm': 40, 'width': 250, 'kw': '薰衣草'},
    # BES-FM-W01-222503000（親膚22gsm*250mm）→ 庫存用 20gsm*260mm 雙壓(福綿) 替代
    'BES-FM-W01-222503000': {'cat': '親膚', 'gsm': 20, 'width': 260, 'brand': '福綿'},
    # BPP-HD-BK03502600700（黑泡泡紋 50gsm）→ 庫存記為 40gsm*260mm*黑泡泡
    'BPP-HD-BK03502600700': {'cat': '外層', 'gsm': 40, 'width': 260, 'kw': '黑泡泡'},
    # BPP-HD-WH02502600700（白泡泡紋 50gsm）→ 庫存記為 40gsm*260mm*白泡泡
    'BPP-HD-WH02502600700': {'cat': '外層', 'gsm': 40, 'width': 260, 'kw': '白泡泡'},
    # BPP-SS-WH01402501500（TN95 4D 冰晶白外層 40*250）
    #   → 實際用料為 40*260白（非 50*250 FFP2 專用布），精確比對 kw='白'
    'BPP-SS-WH01402501500': {'cat': '外層', 'gsm': 40, 'width': 260, 'kw': '白', 'kw_exact': True},
    # BPP-SS-BK01402501500（炫耀黑 40*250）
    #   → 與 BPP-SS-BK01402601500 / BPP-SS250-BLK001-040 共用同一批 40*260黑 庫存
    'BPP-SS-BK01402501500': {'cat': '外層', 'gsm': 40, 'width': 260, 'kw': '黑', 'kw_exact': True},
}

# BPA 包裝盒品號 → 盒子庫存名稱 硬對應表
# 用途：關鍵字比對無法正確識別時，直接指定對應的盒子庫存名稱
# 說明：品名相同或通路商名稱不同但盒子相同時（如 麗德=杏一），由此表強制映射
BPA_BOX_MAP = {
    # ── 玩美系列（平面）── 兩個品號同用一個盒子，庫存鍵='玩美平面'
    # 5/6 進貨 8088 個，尚未打入庫存（待更新盒子庫存後缺料自動消除）
    'BPA-KD2D-DRX029-050': '玩美平面',    # 凱達成人平面-玩美萬用盒-50片
    'BPA-KD2D-DRX031-050': '玩美平面',    # 凱達成人平面-玩美直立式萬用盒
    # ── 玩美3D ──
    'BPA-KD3U-DRX030-030': '玩美3D',      # 凱達成人3D-玩美萬用盒-30片
    # ── 麗德/杏一（同一通路商，不同名稱）──
    'BPA-KD2D-DRX009-050': '杏一灰50',    # 麗德直立彩盒(太空灰) = 杏一灰50
    'BPA-KD2D-BLK002-050': '杏一黑50',    # 凱達成人平面-極致黑 = 麗德/杏一黑50
    # ── 燕麥可可 ──
    'BPA-KD2D-BRN006-050': '可可',         # 凱達成人平面-燕麥可可-50片 = 可可
}

def find_stock(mat_code, mat_name, roll_len=1500):
    """根據 BOM 材料品號 + 品名 回傳 (庫存量捲數, 對應描述)
    roll_len: 每捲米數（從BOM品號末段取得），用於將剩餘米數換算成捲數"""
    code = str(mat_code).strip()

    # ── 耳繩 ──
    if code.startswith('BEL'):
        # 顏色英文→中文對照
        _COLOR_MAP = {
            'WHT': '白', 'WH': '白',
            'BLK': '黑', 'BK': '黑',
            'BLU': '藍',
            'PNK': '粉',
            'GRY': '灰',
            'PUR': '紫',
            'GRN': '綠',
            'YLW': '黃',
            'ORG': '橘',
            'BRN': '棕',
        }
        last_seg = code.split('-')[-1]   # e.g. 'WHT500'
        color_en = re.match(r'([A-Z]+)', last_seg)
        color_en = color_en.group(1) if color_en else ''
        color_zh = _COLOR_MAP.get(color_en, '')

        # ── 綠沅（BEL-LYxxx）：數字色碼 + 顏色雙重確認，避免誤匹配 ──
        # BEL-LY060-BLK001：BLK=黑(色), 001=品號流水號 → 應對到「黑元筋」，不是「冷灰藍001」
        # BEL-LY060-BLU349：BLU=藍, 349=實際色碼 → 應對到「淺藍349」
        if '-LY' in code:
            ly_entries = [e for e in ear_table if e.get('section', '').startswith('綠沅')]
            cn_m = re.search(r'(\d{3,})', last_seg)
            if cn_m:
                cn = cn_m.group(1)
                # 在綠沅庫存中找含該數字的條目，且顏色也一致才採用
                for e in ly_entries:
                    if cn in e['raw'] and color_zh and color_zh in e['code']:
                        return e['qty'], e['raw'][:40]
            # 數字+顏色未能對應 → 以顏色找「{顏色}元筋」（如黑元筋、白元筋）
            if color_zh:
                target = f'{color_zh}元筋'
                for e in ly_entries:
                    if e['code'] == target:
                        return e['qty'], e['raw'][:40]
            return 0, '耳繩待確認'

        # 方法1：顏色碼末段數字精確比對（e.g. 335 → 335珊瑚橘*3）
        # 注意：僅適用於喬煒/一般耳繩，綠沅已在上方處理
        cn_m = re.search(r'(\d{3,})', last_seg)
        if cn_m:
            cn = cn_m.group(1)
            for e in ear_table:
                if cn in e['raw']:
                    return e['qty'], e['raw'][:40]

        # 方法2：從品號 JWxxx 取直徑 + 顏色中文比對（e.g. JW037+WHT→3.7白）
        diam_m = re.search(r'JW(\d{3})', code)
        if diam_m:
            diam = f"{int(diam_m.group(1))/10:.1f}"   # '037' → '3.7'
            if color_zh:
                for e in ear_table:
                    if diam in e['code'] and color_zh in e['code']:
                        return e['qty'], e['raw'][:40]

        # 方法3：從規格取直徑 + 顏色中文比對
        spec_str = spec_lookup.get(code, '')
        diam_from_spec = re.search(r'寬\s*(\d+(?:\.\d+)?)\s*mm', spec_str)
        if diam_from_spec:
            diam = diam_from_spec.group(1)
            if color_zh:
                for e in ear_table:
                    if diam in e['code'] and color_zh in e['code']:
                        return e['qty'], e['raw'][:40]

        return 0, '耳繩待確認'

    # ── 印花外層不織布（BPE-FM-* / BPET-*）：隨訂單叫貨，不在庫存檔
    #    直接回傳 0 / '需叫布'，讓缺料清單顯示「需叫布」提醒
    if code.startswith('BPE') or code.startswith('BPET'):
        return 0, '需叫布'

    # ── 印花外層（BPP-FM-ZA* / BPP-FM-XA*）：有圖案設計，隨訂單叫貨
    if code.startswith('BPP-FM-ZA') or code.startswith('BPP-FM-XA'):
        return 0, '需叫布'

    # ── BJF = 彈性布（3D彈力口罩用，隨訂單叫貨，不在庫存檔）
    if code.startswith('BJF'):
        return 0, '需叫布'

    # ── BAC = 活性碳不織布：直接入現場倉，無主庫存記錄，同印花布顯示需叫布
    if code.startswith('BAC'):
        return 0, '需叫布'

    # ── BPA 包裝盒（從原料類品號取中文品名，比對盒子庫存）──────────────
    if code.startswith('BPA'):
        # 袋裝（TO2D-*）跳過，不追蹤
        if 'TO2D' in code:
            return 0, '不追蹤'
        if not box_inventory:
            return 0, '盒子未追蹤'

        # 優先查硬對應表（比關鍵字比對更可靠）
        if code in BPA_BOX_MAP:
            _mapped_key = BPA_BOX_MAP[code]
            _mapped_qty = box_inventory.get(_mapped_key, 0)
            return int(_mapped_qty), _mapped_key

        _bpa_name = invi_name.get(code, '')
        if not _bpa_name:
            return 0, '盒子待確認'
        # 從 BPA 品名提取關鍵字，比對盒子庫存中文名
        # 品名縮寫替換（TN95 3D → TN3D，TN95 4D → TN4D）
        _BPA_NAME_SUBS = [
            (r'TN95\s*3D', 'TN3D'),
            (r'TN95\s*4D', 'TN4D'),
        ]
        _bpa_name_norm = _bpa_name
        for _pat, _rep in _BPA_NAME_SUBS:
            _bpa_name_norm = re.sub(_pat, _rep, _bpa_name_norm)

        _STOP_KW = {'包裝盒', '達特世', '醫用', '平面', '包裝袋', '一般', '固定', '套', '款', '包裝'}
        # 基本關鍵字（2字以上，不在停用詞中）
        _kws = [w for w in re.split(r'[-\s\(\)（）－＊·,，]+', _bpa_name_norm)
                if w and w not in _STOP_KW and len(w) >= 2]
        # 補入「X片」格式（如 20片, 30片, 50片）—— 2字剛好夠但需特別處理數字片數
        for _pm in re.finditer(r'(\d+)\s*片', _bpa_name):
            _pcs_kw = _pm.group(1) + '片'
            if _pcs_kw not in _kws:
                _kws.append(_pcs_kw)
        # 補入「入」數（如 20入, 30入）→ 盒子庫存多用「片」，也試試「入」→「片」對應
        for _em in re.finditer(r'(\d+)\s*入', _bpa_name_norm):
            _e_kw = _em.group(1) + '片'   # 「入」→「片」嘗試匹配
            if _e_kw not in _kws:
                _kws.append(_e_kw)
        # 補入尺寸代號（XL/L/M/S）—— 單字元也加入以區分 M/L/S 型號
        for _sz in ['XL', 'L', 'M', 'S']:
            if re.search(r'(?<![A-Za-z0-9])' + _sz + r'(?![A-Za-z0-9])', _bpa_name_norm):
                if _sz not in _kws:
                    _kws.append(_sz)

        # ── TN系列（TN3D / TN4D）補入精確關鍵字 ──────────────────────────
        # 盒子庫存格式：TN3D{顏色縮寫}{尺寸}（如TN3D黑L）
        #              TN4D{成人|童}{顏色縮寫}（如TN4D成人白）
        # 品名格式：凱達[成人|兒童]TN4D-[顏色全名]-[片數]
        #          → 被split成一個token「凱達成人TN4D」，需拆出 TN4D + 成人/童 + 顏色縮寫
        _TN_COLOR_ABBR = [
            ('冰晶白', '白'), ('白色', '白'),
            ('炫耀黑', '黑'), ('黑色', '黑'),
            ('都會四色', '都會'), ('都會', '都會'),
            ('胭脂四色', '胭脂'), ('胭脂', '胭脂'),
            ('莫內四色', '莫內'), ('莫內', '莫內'),
            ('莫蘭迪四色', '莫蘭迪'), ('莫蘭迪', '莫蘭迪'),
            ('馬卡龍四色', '馬卡龍'), ('馬卡龍', '馬卡龍'),
        ]
        for _tn in ['TN4D', 'TN3D']:   # TN4D 先判斷（避免含TN4D的名稱被TN3D誤抓）
            if _tn in _bpa_name_norm:
                if _tn not in _kws: _kws.append(_tn)
                # TN4D 盒名有「成人」/「童」區分；TN3D 盒名無此欄位
                if _tn == 'TN4D':
                    if '成人' in _bpa_name_norm and '成人' not in _kws: _kws.append('成人')
                    if '兒童' in _bpa_name_norm and '童'  not in _kws: _kws.append('童')
                # 顏色縮寫（只補第一個命中的）
                for _fc, _ac in _TN_COLOR_ABBR:
                    if _fc in _bpa_name_norm:
                        if _ac not in _kws: _kws.append(_ac)
                        break
                break   # 只處理一個 TN 系列

        # 計分：每個關鍵字命中 +1；完全相同的關鍵字得雙倍分（精確度獎勵）
        best_key, best_score = None, 0
        for _bk in box_inventory:
            score = sum(2 if kw == _bk else 1 for kw in _kws if kw in _bk)
            # 同分時選較短的（避免選到「萬用盒」這種包含太多品名的條目）
            if score > best_score or (score == best_score and best_key and len(_bk) < len(best_key)):
                best_score, best_key = score, _bk
        if best_key and best_score > 0:
            return int(box_inventory[best_key]), best_key
        # fallback：找含任一關鍵字、且名稱最短的盒子（避免過長的萬用盒搶走位置）
        candidates = [(len(_bk), _bk) for _bk in box_inventory if any(kw in _bk for kw in _kws)]
        if candidates:
            return int(box_inventory[min(candidates)[1]]), min(candidates)[1]
        return 0, f'盒子待確認({_bpa_name[:20]})'

    # ── 不追蹤的品碼前綴（包材、鼻線、PE膜等）──
    # 注意：BPE 已移出（BPE-FM-* 是印花布，非PE膜）；PE膜用 FPE 前綴
    # BPA 已移出（包裝盒，需追蹤）；BBO/BBX 若無庫存檔則不追蹤
    skip_prefix = ('BNW', 'FPEB', 'FPE', 'BBO', 'BBX',
                   'BLB', 'BNS', 'BPK', 'BST', 'BTK', 'BWR')
    if any(code.startswith(p) for p in skip_prefix):
        return 0, '不追蹤'

    # ── 特定品號替代庫存（BOM_OVERRIDES）──
    if code in BOM_OVERRIDES:
        ov  = BOM_OVERRIDES[code]
        rl  = roll_len if roll_len > 0 else 1500
        ov_cat = ov.get('cat', '')
        if ov_cat == '外層':
            # kw_exact=True 時用精確比對（避免'白'命中'白泡泡'）
            if ov.get('kw_exact', False):
                hits = [it for it in inv_table
                        if it['cat'] == '外層'
                        and it['gsm'] == ov['gsm']
                        and it['width'] == ov['width']
                        and it['kw'] == ov['kw']]
            else:
                hits = [it for it in inv_table
                        if it['cat'] == '外層'
                        and it['gsm'] == ov['gsm']
                        and it['width'] == ov['width']
                        and ov.get('kw', '') in it['kw']]
            if hits:
                total = sum(it.get('qty_rolls', it['qty']) + it.get('qty_m', 0) / rl
                            for it in hits)
                return round(total, 2), hits[0]['raw'][:40]
        elif ov_cat == '親膚':
            tot, raw = 0, ''
            for it in inv_table:
                if it['cat'] != '親膚': continue
                if it['gsm'] != ov['gsm']: continue
                if it['width'] != ov['width']: continue
                if it['kw'] != ov.get('brand', ''): continue
                tot += it.get('qty_rolls', it['qty']) + it.get('qty_m', 0) / rl
                if not raw: raw = it['raw'][:40]
            if raw:
                return round(tot, 2), raw
        return 0, '待確認'

    # ── 布料 ──
    if 'BES-FM' in code:
        cat, brand = '親膚', '福綿'
    elif 'BES-JG' in code:
        cat, brand = '親膚', '嘉谷'
    elif code.startswith('BMB'):
        cat = '熔噴'
        if   'MEDG' in code: brand = '醫灰'
        elif 'MEDW' in code: brand = '醫白'
        elif 'H12'  in code: brand = 'H12'
        elif 'H11'  in code: brand = 'H11'
        else:                brand = '醫白'
    elif code.startswith('BPP'):
        cat, brand = '外層', ''
    elif code.startswith('BES'):
        cat, brand = '親膚', '福綿'  # 其他BES格式預設福綿
    else:
        return 0, '待確認'

    # ── 從品號中提取 gsm + width ──
    # 統一方式：在去掉連字符的品號中，找「2位gsm + 已知width(3位)」
    # 例：BES-FM-W01-201753000  → BESFMW01201753000  → 20+175
    #     BMB-MG-MEDW201751500  → BMBMGMEDW201751500 → 20+175
    #     BPP-SS-BL01301751500  → BPPSSBL01301751500 → 30+175
    # 特殊格式A: BES-FM220-WHT025 → width=220 在品牌段, gsm=025 在末段
    # 特殊格式B: BPP-SS175-BLU001-020 → width=175 在類型段, gsm=020 在末段
    KNOWN_WIDTHS = r'170|175|190|195|200|220|230|240|250|260|270|280'
    code_nd = code.replace('-', '')
    m_dims = re.search(rf'([1-9]\d)({KNOWN_WIDTHS})', code_nd)
    gsm, width = None, None
    if m_dims:
        gsm   = int(m_dims.group(1))
        width = int(m_dims.group(2))
    else:
        # 嘗試特殊格式：品號段中含 WIDTH，gsm 在末段（如 025=25g, 040=40g）
        segs = code.split('-')
        for seg in segs:
            mw = re.search(rf'({KNOWN_WIDTHS})', seg)
            if mw:
                width = int(mw.group(1))
                break
        # gsm 從末段取（3位，如 025 → 25）
        if width:
            m_gsm_tail = re.match(r'^0?(\d{1,2})$', segs[-1])
            if m_gsm_tail:
                gsm = int(m_gsm_tail.group(1))
    if not width:
        return 0, '待確認'

    # 顏色縮寫 → 中文關鍵字（從品號字母部分提取）
    COLOR_MAP = {
        'BL': '藍', 'BLU': '藍',
        'WH': '白', 'WHT': '白',
        'BK': '黑', 'BLK': '黑',
        'GY': '灰', 'GRY': '灰',
        'PK': '粉', 'PNK': '粉',
        'BR': '棕', 'BRN': '棕',
        'GN': '綠', 'GRN': '綠',
        'PR': '紫', 'PUR': '紫',
        'YW': '黃', 'YLW': '黃',
        'OR': '橘', 'ORG': '橘',
        'RD': '紅', 'MGY': '灰',
        'MBR': '棕',
    }
    color_kw = None
    if cat == '外層':
        # 從品號字母段找顏色縮寫（取 gsm 前面那一段）
        prefix = code_nd[:m_dims.start()] if m_dims else code_nd
        for abbr in sorted(COLOR_MAP.keys(), key=len, reverse=True):  # 長的優先
            if abbr in prefix:
                color_kw = COLOR_MAP[abbr]; break
        if not color_kw:
            for c in ['黑', '白', '藍', '灰', '粉', '紫', '綠', '橘', '茶', '奶', '紅', '棕']:
                if c in str(mat_name):
                    color_kw = c; break

    # ── 外層：精確比對（避免把碧藍+天空藍+莫B藍...全部加在一起）──
    if cat == '外層':
        pool_exact = [it for it in inv_table
                      if it['cat'] == cat
                      and (gsm is None or it['gsm'] == gsm)
                      and it['width'] == width]
        # 放寬 GSM 的替代品池（如 35*200→30*200，佳得利035→30*200庫存）
        pool_relax = [it for it in inv_table
                      if it['cat'] == cat
                      and it['width'] == width]
        if not pool_relax:
            return 0, '待確認'

        def _inv_qty(it):
            """將庫存品項換算成捲數（完整捲 + 剩餘米數÷每捲米數）"""
            rolls = it.get('qty_rolls', it['qty'])
            extra_m = it.get('qty_m', 0)
            rl = roll_len if roll_len > 0 else 1500
            return rolls + extra_m / rl

        def best_kw_match(p, name_str):
            """從指定 pool 中找品名最長匹配的庫存品項"""
            hits = [it for it in p if it['kw'] and it['kw'] in name_str]
            if not hits:
                return None
            best_len = max(len(it['kw']) for it in hits)
            return [it for it in hits if len(it['kw']) == best_len]

        def _try_name(p):
            """方法1+2：BOM品名 / INVI02品名比對，較精確，優先跑"""
            name_str = str(mat_name).strip()
            if name_str:
                best = best_kw_match(p, name_str)
                if best:
                    return round(sum(_inv_qty(it) for it in best), 2), best[0]['raw'][:40]
            invi_nm = invi_name.get(code, '')
            if invi_nm:
                best = best_kw_match(p, invi_nm)
                if best:
                    return round(sum(_inv_qty(it) for it in best), 2), best[0]['raw'][:40]
            return None

        def _try_color(p):
            """方法3+4：color_kw 比對，作為次要 fallback"""
            if color_kw:
                exact = [it for it in p if it['kw'] == color_kw]
                if exact:
                    return round(sum(_inv_qty(it) for it in exact), 2), exact[0]['raw'][:40]
            if color_kw:
                for it in p:
                    # 方法4：子字串匹配，但跳過複合色（如「灰紫藍」含多個顏色字）
                    if color_kw in it['kw'] and not _is_compound_color(it['kw'], color_kw):
                        return round(_inv_qty(it), 2), it['raw'][:40]
            return None

        # 步驟1：BOM品名/INVI02 先跨兩個 pool 比對（最精確，如「櫻花粉」能對到正確品項）
        result = _try_name(pool_exact) if pool_exact else None
        if not result:
            result = _try_name(pool_relax)
        if result:
            return result

        # 步驟2：顏色縮寫比對（在精確 GSM pool 中）
        result = _try_color(pool_exact) if pool_exact else None
        if result:
            return result
        # 步驟3：顏色縮寫比對（放寬 GSM pool）
        result = _try_color(pool_relax)
        if result:
            return result

        # 寬度替代（如 195→200, 250→260）
        sub_w = WIDTH_SUBS.get(width)
        if sub_w:
            pool_sub = [it for it in inv_table if it['cat'] == cat and it['width'] == sub_w]
            result = _try_name(pool_sub) or _try_color(pool_sub)
            if result:
                return result

        return 0, '待確認'

    # ── 親膚 / 熔噴：brand 精確比對 ──
    def _match_skin_melt(w):
        tot, raw = 0, ''
        rl = roll_len if roll_len > 0 else 1500
        for item in inv_table:
            if item['cat'] != cat:
                continue
            if gsm is not None and item['gsm'] != gsm:
                continue
            if item['width'] != w:
                continue
            if item['kw'] != brand:
                continue
            # 完整捲 + 剩餘米數換算
            rolls_f = item.get('qty_rolls', item['qty'])
            extra_m = item.get('qty_m', 0)
            tot += rolls_f + extra_m / rl
            if not raw:
                raw = item['raw'][:40]
        return round(tot, 2), raw

    total, matched_raw = _match_skin_melt(width)
    if matched_raw:
        return total, matched_raw

    # 寬度替代（如親膚250→260）
    sub_w = WIDTH_SUBS.get(width)
    if sub_w:
        total, matched_raw = _match_skin_melt(sub_w)
        if matched_raw:
            return total, matched_raw

    return 0, '待確認'

# ─── 3. 讀取 BOM ─────────────────────────────────────────

def roll_len(spec):
    """每捲幾米"""
    m = re.search(r'(\d+)\s*m/[捲Rr]', str(spec), re.I)
    return float(m.group(1)) if m else 0

def ear_density(spec):
    """每 kg 幾米"""
    m = re.search(r'(\d+)\s*m/kg', str(spec), re.I)
    return float(m.group(1)) if m else 830

# ── 3-A. 主BOM（原物料用途清單）：建立規格查找表 + 備用BOM ──
spec_lookup  = {}   # 元件品號 → 規格字串
master_bom   = None # 備用BOM DataFrame

if master_path:
    print(f"\n讀取主BOM: {os.path.basename(master_path)}")
    mb_raw = pd.read_excel(master_path, sheet_name=0)
    # 清理欄位名稱（移除全形空格 + 一般空白）
    mb_raw.columns = [re.sub(r'\s+', '', str(c)) for c in mb_raw.columns]
    print(f"  主BOM 欄位: {mb_raw.columns.tolist()}")

    # 採購件行（主件品號為空）→ 取規格
    mask_purch = mb_raw['主件品號'].isna() | (mb_raw['主件品號'].astype(str).str.strip() == '')
    for _, row in mb_raw[mask_purch].iterrows():
        code = str(row['元件品號']).strip()
        spec = str(row.get('規格', '')).strip()
        if code and spec and spec.lower() != 'nan':
            spec_lookup[code] = spec

    # 自製件行 → 備用BOM（製令總表找不到的品號可從此補用）
    mb_mfg = mb_raw[~mask_purch].copy()
    mb_mfg['損耗率%']  = pd.to_numeric(mb_mfg['損耗率%'],  errors='coerce').fillna(0)
    mb_mfg['組成用量'] = pd.to_numeric(mb_mfg['組成用量'], errors='coerce').fillna(0)
    master_bom = mb_mfg
    print(f"  規格查找表: {len(spec_lookup)} 筆  |  備用BOM: {len(master_bom)} 筆")
    # 印出耳繩規格（供確認 m/kg）
    ear_specs = {k: v for k, v in spec_lookup.items() if k.startswith('BEL')}
    if ear_specs:
        print("  耳繩規格:")
        for k, v in list(ear_specs.items())[:10]:
            print(f"    {k:30}  {v}")
else:
    print("  ⚠ 未找到原物料用途清單，耳繩密度使用預設 830m/kg")

# ── 3-B. 每日製令BOM ──────────────────────────────────────
print(f"\n讀取製令BOM: {os.path.basename(bom_path)}")
bom_df = pd.read_excel(bom_path, sheet_name="BOM資料")
print(f"  製令BOM 欄位: {bom_df.columns.tolist()}")

bom = bom_df[bom_df['主件品號'].notna()].copy()
bom['損耗率%']  = pd.to_numeric(bom['損耗率%'],  errors='coerce').fillna(0)
bom['組成用量'] = pd.to_numeric(bom['組成用量'], errors='coerce').fillna(0)

# ─── 4. 讀取訂單 ─────────────────────────────────────────
print(f"讀取訂單: {os.path.basename(order_path)}")
try:
    _ord_tmp = os.path.join(tempfile.gettempdir(), 'ord_tmp_' + os.path.basename(order_path))
    try:
        shutil.copy2(order_path, _ord_tmp)
        _ord_load = _ord_tmp
    except Exception:
        _ord_load = order_path
    xl_o = pd.ExcelFile(_ord_load)
    print(f"  工作表: {xl_o.sheet_names}")
    order_df = None
    # 優先使用名稱完全為「總表」的工作表，其次含「總表」，最後第一個
    sheet_priority = (
        [sn for sn in xl_o.sheet_names if str(sn).strip() == '總表'] or
        [sn for sn in xl_o.sheet_names if '總表' in str(sn)] or
        [xl_o.sheet_names[0]]
    )
    sn = sheet_priority[0]
    order_df = xl_o.parse(sn)
    print(f"  使用: [{sn}]  欄位: {order_df.columns.tolist()}")
except Exception as e:
    print(f"  ❌ 訂單讀取失敗: {e}"); sys.exit(1)

order_df['生產量'] = pd.to_numeric(order_df.get('生產量', pd.Series(dtype=float)),
                                    errors='coerce').fillna(0)
# 生產量為0但差異>0 → 尚未開製令的訂單，以「差異」欄當作生產量
if '差異' in order_df.columns:
    差異_s = pd.to_numeric(order_df['差異'], errors='coerce').fillna(0)
    order_df['生產量'] = order_df['生產量'].where(order_df['生產量'] > 0, 差異_s)
    n_diff = ((order_df['生產量'] > 0) & (差異_s > 0) & (pd.to_numeric(order_df.get('生產量', 0), errors='coerce').fillna(0) == 0)).sum()
active = order_df[order_df['生產量'] > 0].dropna(subset=['品號']).copy()

# ── 排除已備料/已完工訂單：
#   入庫有日期 = 已完工入庫（排除）
#   領料有值且不含「半」字 = 材料已全數領出（排除）
#   領料含「半」字（如「領一半」）= 尚未完全備料，仍需計算缺料（保留）
def _col_filled(df, col):
    """判斷欄位是否已填寫（含值且不含「半」字才算完整填寫）"""
    if col not in df.columns: return pd.Series(False, index=df.index)
    filled  = df[col].notna() & ~(df[col].astype(str).str.strip().isin(['', 'nan', 'NaT']))
    partial = df[col].astype(str).str.contains('半', na=False)  # 「領一半」不算完整
    return filled & ~partial

before = len(active)
入庫_mask = _col_filled(active, '入庫')
領料_mask = _col_filled(active, '領料')
done_mask = 入庫_mask | 領料_mask
n_入庫 = 入庫_mask.sum()
n_領料 = (~入庫_mask & 領料_mask).sum()
if done_mask.any():
    print(f"  ⚙ 排除已入庫: {n_入庫} 筆  排除已完整領料: {n_領料} 筆")
active = active[~done_mask].copy()
print(f"  排除後: {before} → {len(active)} 筆")

# ── 須領料用：排除完工後、去重複前，保留全部訂單行（同批號不合併）──
active_for_pick = active.copy()

# ── 去重複：
#   有製令單號 → 同製令+品號視為同一生產批次，只計算一次
#   無製令單號 → 不同客戶同品號訂單，全部保留（各自計算需求）
dup_col = None
for c in ['製令單號', '批號']:
    if c in active.columns:
        dup_col = c; break
if dup_col:
    before = len(active)
    _has_order = (active[dup_col].notna()
                  & ~active[dup_col].astype(str).str.strip().isin(['', 'nan', 'NaT']))
    grp_with = (active[_has_order]
                .sort_values(dup_col)
                .drop_duplicates(subset=['品號', dup_col], keep='first'))
    grp_none = active[~_has_order]   # 無製令：保留但先去除完全重複的序列
    # 無製令訂單：若「序列」欄位存在，相同序列只保留一筆（避免總表中重複列）
    if '序列' in grp_none.columns:
        n_before_seq = len(grp_none)
        grp_none = grp_none.drop_duplicates(subset=['序列', '品號'], keep='first')
        n_seq_removed = n_before_seq - len(grp_none)
        if n_seq_removed:
            print(f"  ⚙ 無製令去重複序列: 移除 {n_seq_removed} 筆重複序列")
    active = pd.concat([grp_with, grp_none], ignore_index=True)
    removed = before - len(active)
    if removed:
        print(f"  ⚙ 去重複（同{dup_col}+品號）: 移除 {removed} 筆重複; 無製令訂單保留 {len(grp_none)} 筆")
print(f"有效訂單: {len(active)} 筆（含無製令 {(~_has_order if dup_col else pd.Series(False, index=active.index)).sum()} 筆）")

# ─── 5. 計算需求明細 ───────────────────────────────────────
details = []
no_bom  = []

for _, o in active.iterrows():
    pno        = str(o['品號']).strip()
    qty        = float(o['生產量'])
    batch      = o.get('批號', '')
    pname      = o.get('品名', pno)
    order_unit = str(o.get('單位', '盒')).strip()  # 盒 / 袋 / 片

    # ── 領一半：只剩一半料尚未領，需求量 ÷ 2 ──
    _ling = str(o.get('領料', '')).strip()
    if '半' in _ling:
        qty = qty * 0.5

    # ── 袋裝換算：5入/袋 → 標準50入/盒 等效盒數（÷10）
    # 算法：從品名取每袋入數，除以標準盒入數(50)，換算成等效盒數
    # 例：20000袋(5入) → 20000×5÷50 = 2000 等效盒
    if order_unit == '袋':
        _m_bag = re.search(r'(\d+)\s*入', str(pname))
        _bag_pcs = int(_m_bag.group(1)) if _m_bag else 5   # 預設5入/袋
        qty_box_equiv = qty * _bag_pcs / 50                 # 換算成50入/盒的等效盒數
    else:
        qty_box_equiv = qty                                  # 盒裝直接用

    # 先查製令BOM，找不到再查主BOM
    parts = bom[bom['主件品號'].astype(str).str.strip() == pno]
    bom_source = '製令'
    if parts.empty and master_bom is not None:
        parts = master_bom[master_bom['主件品號'].astype(str).str.strip() == pno]
        bom_source = '主BOM'
    if parts.empty:
        no_bom.append(pno); continue
    if bom_source == '主BOM':
        print(f"  [{pno}] 製令BOM無資料，改用主BOM備用")

    for _, p in parts.iterrows():
        mat_code = str(p['元件品號']).strip()

        # ── 產品特定耗材替換（BOM已確認改料，尚未更新系統）──
        # (成品品號, 原BOM料號) → 替換料號
        PRODUCT_MAT_SUBS = {
            ('AMDMSK-2D1-BLU101-50', 'BEL-JW035-WHT737'): 'BEL-JW031-WHT373',
            ('AMDMSK-2D1-PUR115-50', 'BEL-JW035-WHT737'): 'BEL-JW031-WHT373',
        }
        sub_key = (pno, mat_code)
        if sub_key in PRODUCT_MAT_SUBS:
            mat_code = PRODUCT_MAT_SUBS[sub_key]

        mat_name = str(p.get('品名', '')) if '品名' in p.index else ''
        usage    = float(p['組成用量'])
        loss_pct = float(p['損耗率%'])

        # 規格：優先從主BOM規格查找表取（更準確），其次用製令BOM的規格欄
        spec = (spec_lookup.get(mat_code)
                or (str(p.get('規格', '')) if '規格' in p.index else ''))

        demand_m  = qty * usage * (1 + loss_pct / 100)
        mtype     = mat_code_type(mat_code)
        rl        = roll_len(spec)
        ed        = ear_density(spec)

        # ★ 若規格欄沒有 m/捲，從品號末段數字推算捲長
        # BES-FM-W01-201753000 末段3000 → 3000m/捲
        # BMB-MG-MEDW201751500 末段1500 → 1500m/捲
        # BPP-SS-BL01301751500 末段1500 → 1500m/捲
        # BPP-SS175-BLU001-020  末段020=gsm，倒數第二段001→捲長無法取，預設1500
        # BES-FM220-WHT025 末段025=gsm，預設3000
        if rl == 0 and mtype != '耳繩':
            m_rl = re.search(r'(\d{3,4})$', mat_code.replace('-', ''))
            if m_rl:
                cand = int(m_rl.group(1))
                if 500 <= cand <= 9999:
                    rl = cand
                elif cand <= 100:
                    # 末段是gsm（如020=20g），取倒數第二段找捲長
                    segs = mat_code.split('-')
                    if len(segs) >= 2:
                        m_rl2 = re.search(r'(\d{3,4})$', segs[-2])
                        if m_rl2:
                            cand2 = int(m_rl2.group(1))
                            if 500 <= cand2 <= 9999:
                                rl = cand2
            # 以上方法都找不到捲長（如品號末段為 H11/H12 等字母）→ 用預設值
            if rl == 0:
                # 品號含BES/BMB預設3000，BPP預設1500
                rl = 3000 if mtype in ('親膚',) else 1500

        # ── 外層布（BPP 品號）改用換算表計算需求 ─────────────────────────
        # 平面：盒數 ÷ 160 = 捲數（進位），需求米數 = 捲數 × 每捲米數
        # 3D/醫用等：盒數 × 每盒片數 ÷ 每100M片數 × 100 = 需求米數
        #   每100M片數：L=680, M=730, S=790
        #   每盒片數：從品名「30入」取，預設30
        # 多色產品（多個BPP元件）：各元件平分總需求
        if mat_code.startswith('BPP') and rl > 0:
            # 計算此訂單品號的 BPP 元件數量（多色分攤用）
            n_bpp = max(1, int((parts['元件品號'].astype(str).str.strip()
                                .str.startswith('BPP')).sum()))
            # 使用等效盒數（袋裝已換算，盒裝不變）
            qty_per_bpp = qty_box_equiv / n_bpp

            if '平面' in pname:
                # 平面公式：等效盒數 ÷ 160，無條件進位
                demand_rolls = math.ceil(qty_per_bpp / 160)
                demand_m = demand_rolls * rl
            else:
                # 3D/醫用：偵測尺寸，取每100M片數
                # ⚠️ XL 必須在 L 之前檢查，否則 XL 會被誤判為 L
                _size_map = {'XL': 630, 'L': 680, 'M': 730, 'S': 790}
                _size_ch = None
                for _sz in ['XL', 'L', 'M', 'S']:
                    if (re.search(r'(?<![A-Za-z])' + _sz + r'(?![A-Za-z])', pname)
                            or pname.endswith(f'-{_sz}')
                            or f'({_sz})' in pname
                            or f'-{_sz}(' in pname):
                        _size_ch = _sz
                        break
                _ppm = _size_map.get(_size_ch, 730)  # 找不到尺寸預設 M=730

                # ⚠️ TN95 修正（2026-05-14）：
                #   TN95 口罩展開面積比普通 3D 大，3D 與 4D 用量相同
                #   BOM驗算：
                #     成人 L/XL  單色/MIX → 417片/100m（0.24m/片）
                #     兒童/幼幼 M/S 單色   → 541片/100m（0.185m/片）
                #     兒童/幼幼 M/S MIX多色 → 500片/100m（0.20m/片）
                #   品號：AMDMSK-T3*（TN95 3D）或 AMDMSK-TN*（TN95 4D）
                if (pno.startswith('AMDMSK-TN') or pno.startswith('AMDMSK-T3')
                        or 'TN95' in pname):
                    if _size_ch in ('M', 'S'):
                        _ppm = 500 if n_bpp > 1 else 541  # MIX多色=500，單色=541
                    else:
                        _ppm = 417   # 成人 L/XL（單色或MIX相同）

                # 每包片數：從品名「30入」取（盒/袋通用，直接用品名的入數）
                _m_ppb = re.search(r'(\d+)\s*入', pname)
                _ppb = int(_m_ppb.group(1)) if _m_ppb else 30

                # 需求米數（含損耗）：total片數 / 每100M片數 × 100
                # ⚠️ 袋裝修正（2026-05-05）：
                #   qty_per_bpp 已換算成「等效50入盒數」，故乘以 50（非袋入數）
                #   盒裝：qty_box_equiv = qty → qty × _ppb / _ppm × 100 ✓
                #   袋裝：qty_box_equiv = qty × bag_pcs/50 → 需再乘50才能還原總片數
                if order_unit == '袋':
                    # 袋裝：total片數 = qty × _bag_pcs，不需再乘bag_pcs
                    total_pcs_per_bpp = qty * _bag_pcs / n_bpp
                    demand_m = total_pcs_per_bpp / _ppm * 100 * (1 + loss_pct / 100)
                else:
                    demand_m = qty_per_bpp * _ppb / _ppm * 100 * (1 + loss_pct / 100)

        if mtype == '耳繩':
            need_unit  = round(demand_m / ed, 2) if ed else demand_m
            unit_label = 'kg'
        elif mtype == '盒子':
            # 盒子：BOM用量=1個/盒，demand_m 即等於訂單盒數（不除捲長）
            need_unit  = round(demand_m, 0)
            unit_label = '個'
        elif rl > 0:
            need_unit  = round(demand_m / rl, 2)
            unit_label = '捲'
        else:
            need_unit  = round(demand_m, 2)
            unit_label = 'm'

        # 每捲/kg量（顯示用）
        if mtype == '耳繩':
            per_unit = ed   # m/kg
        elif mtype == '盒子':
            per_unit = 1    # 1個/個（盒子不換算捲數）
        elif rl > 0:
            per_unit = rl   # m/捲
        else:
            per_unit = 0

        details.append({
            '訂單批號':    batch,
            '品號':        pno,
            '品名':        pname,
            '生產量':      qty,
            '材料類型':    mtype,
            '材料品號':    mat_code,
            '材料品名':    mat_name,
            '需求米數(m)': round(demand_m, 2),
            '每捲/kg量':   per_unit,
            '需備量':      need_unit,
            '單位':        unit_label,
        })

if no_bom:
    print(f"  ⚠ 以下品號在 BOM 中找不到: {no_bom}")
print(f"材料需求明細: {len(details)} 筆")

if not details:
    print("⚠ 無任何需求明細，請確認品號與 BOM 對應"); sys.exit(1)

df_det = pd.DataFrame(details)

# ── 將通路資訊合併回 df_det（供大樹獨立缺料使用）
if '通路' in active.columns:
    _pno_ch = dict(zip(active['品號'].astype(str).str.strip(),
                       active['通路'].astype(str).str.strip()))
    df_det['通路'] = df_det['品號'].map(_pno_ch).fillna('')
else:
    df_det['通路'] = ''

# ─── 6. 彙總 + 查庫存 ─────────────────────────────────────
# ★ 只用「材料品號」當 key，確保同一品號只有一行
df_sum = (df_det
          .groupby(['材料品號'], as_index=False)
          .agg(
              材料類型  = ('材料類型',    'first'),
              材料品名  = ('材料品名',    'first'),
              單位      = ('單位',        'first'),
              總需備量  = ('需備量',      'sum'),
              每捲米數  = ('每捲/kg量',   'first'),  # 用於庫存剩餘米數換算
              **{'總需求(m)': ('需求米數(m)', 'sum')}
          ))

# 查庫存（每品號只查一次）
stock_cache = {}
for _, row in df_sum.iterrows():
    k = row['材料品號']
    if k not in stock_cache:
        rl = int(row.get('每捲米數', 1500) or 1500)
        stock_cache[k] = find_stock(row['材料品號'], row['材料品名'], roll_len=rl)

df_sum['庫存量']   = df_sum['材料品號'].map(lambda k: stock_cache[k][0])
df_sum['庫存來源'] = df_sum['材料品號'].map(lambda k: stock_cache[k][1])
df_sum['淨缺量']   = (df_sum['總需備量'] - df_sum['庫存量']).clip(lower=0).round(2)
df_sum['狀態']     = df_sum['淨缺量'].apply(lambda x: '✓ 充足' if x == 0 else '⚠ 缺料')

# ── 共享庫存修正：多個材料品號對應同一實體庫存時，合併計算缺量 ──────
# 例：BPP-SS-BK01302001500（酷黑）與 BPP-HD-BK03302001500（暗夜黑）
# 都對應 30*200黑*3（同 3 捲），不能各自獨立扣庫存
_SKIP_DESCS = {'不追蹤', '待確認', '需叫布', '耳繩待確認', ''}
_desc_to_codes = {}
for _k, (_qty, _desc) in stock_cache.items():
    _d = str(_desc).strip()
    if _d in _SKIP_DESCS or _qty <= 0:
        continue
    if _k in df_sum['材料品號'].values:
        _desc_to_codes.setdefault(_d, []).append(_k)

for _desc, _codes in _desc_to_codes.items():
    if len(_codes) < 2:
        continue
    _mask = df_sum['材料品號'].isin(_codes)
    _shared_stock  = stock_cache[_codes[0]][0]
    _total_demand  = df_sum.loc[_mask, '總需備量'].sum()
    _group_short   = max(0.0, round(_total_demand - _shared_stock, 2))
    # 按需求比例分配缺量到各材料品號
    for _c in _codes:
        _cm = df_sum['材料品號'] == _c
        _cd = float(df_sum.loc[_cm, '總需備量'].values[0])
        _cs = round(_group_short * _cd / _total_demand, 2) if _total_demand > 0 else 0.0
        df_sum.loc[_cm, '淨缺量'] = _cs
    df_sum.loc[_mask, '庫存來源'] = df_sum.loc[_mask, '庫存來源'].apply(
        lambda x: x + '【共用】' if '【共用】' not in str(x) else x)
    print(f"  共享庫存修正 [{_desc}]: {_codes}")
    print(f"    合併需求={round(_total_demand,2)}, 庫存={_shared_stock}, 總缺={_group_short}")

df_sum['狀態'] = df_sum['淨缺量'].apply(lambda x: '✓ 充足' if x == 0 else '⚠ 缺料')

# 不追蹤的品碼不列入缺料清單；「需叫布」印花外層一律列入（庫存恆為0）
shortage = df_sum[(df_sum['淨缺量'] > 0) & (df_sum['庫存來源'] != '不追蹤')].copy()

# ── 各缺料材料 → 涉及訂單生產量（盒）──
# df_det['生產量'] 已套用領一半 ÷ 2，每筆為 (品號, 材料品號, 生產量)
# 同一品號對同一材料只計算一次，避免因多個BOM元件重複加總
_det_uniq = df_det[['品號','材料品號','生產量']].drop_duplicates(subset=['品號','材料品號'])
mat_order_boxes = (
    _det_uniq[_det_uniq['材料品號'].isin(shortage['材料品號'])]
    .groupby('材料品號')['生產量'].sum()
    .round(0).astype(int)
    .to_dict()
)
mat_order_count = (
    _det_uniq[_det_uniq['材料品號'].isin(shortage['材料品號'])]
    .groupby('材料品號')['品號'].nunique()
    .to_dict()
)
# ── 大樹訂單獨立缺料計算 ────────────────────────────────
_datsu_pnos = set(df_det[df_det['通路'] == '大樹']['品號'])
if _datsu_pnos:
    df_det_ds = df_det[df_det['通路'] == '大樹'].copy()
    df_sum_ds = (df_det_ds
                 .groupby(['材料品號'], as_index=False)
                 .agg(材料類型=('材料類型', 'first'),
                      材料品名=('材料品名', 'first'),
                      單位    =('單位',     'first'),
                      總需備量=('需備量',   'sum'),
                      每捲米數=('每捲/kg量','first')))
    df_sum_ds['庫存量']   = df_sum_ds['材料品號'].map(lambda k: stock_cache.get(k,(0,''))[0])
    df_sum_ds['庫存來源'] = df_sum_ds['材料品號'].map(lambda k: stock_cache.get(k,(0,''))[1])
    df_sum_ds['淨缺量']   = (df_sum_ds['總需備量'] - df_sum_ds['庫存量']).clip(lower=0).round(2)
    # 大樹共享庫存修正
    for _desc, _codes in _desc_to_codes.items():
        if len(_codes) < 2:
            continue
        _mask = df_sum_ds['材料品號'].isin(_codes)
        if not _mask.any():
            continue
        _shared_stock = stock_cache[_codes[0]][0]
        _total_demand = df_sum_ds.loc[_mask, '總需備量'].sum()
        _group_short  = max(0.0, round(_total_demand - _shared_stock, 2))
        for _c in _codes:
            _cm = df_sum_ds['材料品號'] == _c
            if not _cm.any():
                continue
            _cd = float(df_sum_ds.loc[_cm, '總需備量'].values[0])
            _cs = round(_group_short * _cd / _total_demand, 2) if _total_demand > 0 else 0.0
            df_sum_ds.loc[_cm, '淨缺量'] = _cs
        df_sum_ds.loc[_mask, '庫存來源'] = df_sum_ds.loc[_mask, '庫存來源'].apply(
            lambda x: x + '【共用】' if '【共用】' not in str(x) else x)
    df_sum_ds['狀態'] = df_sum_ds.get('淨缺量', pd.Series(dtype=float)).apply(
        lambda x: '✓ 充足' if x == 0 else '⚠ 缺料') if '淨缺量' in df_sum_ds.columns else None
    shortage_ds = df_sum_ds[(df_sum_ds['淨缺量'] > 0) & (df_sum_ds['庫存來源'] != '不追蹤')].copy()
    _ds_uniq = df_det_ds[['品號','材料品號','生產量']].drop_duplicates(subset=['品號','材料品號'])
    ds_order_boxes = (_ds_uniq[_ds_uniq['材料品號'].isin(shortage_ds['材料品號'])]
                      .groupby('材料品號')['生產量'].sum().round(0).astype(int).to_dict())
    ds_order_count = (_ds_uniq[_ds_uniq['材料品號'].isin(shortage_ds['材料品號'])]
                      .groupby('材料品號')['品號'].nunique().to_dict())
    print(f"大樹訂單: {len(_datsu_pnos)} 品號  缺料: {len(shortage_ds)} 項")
else:
    df_det_ds = pd.DataFrame(); df_sum_ds = pd.DataFrame()
    shortage_ds = pd.DataFrame(); ds_order_boxes = {}; ds_order_count = {}
    print("大樹訂單: 本次無大樹訂單")

# ── 排除大樹後的缺料計算（只看非大樹訂單）─────────────────
df_det_nds = df_det[df_det['通路'] != '大樹'].copy()
if not df_det_nds.empty:
    df_sum_nds = (df_det_nds
                  .groupby(['材料品號'], as_index=False)
                  .agg(材料類型=('材料類型', 'first'),
                       材料品名=('材料品名', 'first'),
                       單位    =('單位',     'first'),
                       總需備量=('需備量',   'sum'),
                       每捲米數=('每捲/kg量','first')))
    df_sum_nds['庫存量']   = df_sum_nds['材料品號'].map(lambda k: stock_cache.get(k,(0,''))[0])
    df_sum_nds['庫存來源'] = df_sum_nds['材料品號'].map(lambda k: stock_cache.get(k,(0,''))[1])
    df_sum_nds['淨缺量']   = (df_sum_nds['總需備量'] - df_sum_nds['庫存量']).clip(lower=0).round(2)
    # 排除大樹共享庫存修正
    for _desc, _codes in _desc_to_codes.items():
        if len(_codes) < 2:
            continue
        _mask = df_sum_nds['材料品號'].isin(_codes)
        if not _mask.any():
            continue
        _shared_stock = stock_cache[_codes[0]][0]
        _total_demand = df_sum_nds.loc[_mask, '總需備量'].sum()
        _group_short  = max(0.0, round(_total_demand - _shared_stock, 2))
        for _c in _codes:
            _cm = df_sum_nds['材料品號'] == _c
            if not _cm.any():
                continue
            _cd = float(df_sum_nds.loc[_cm, '總需備量'].values[0])
            _cs = round(_group_short * _cd / _total_demand, 2) if _total_demand > 0 else 0.0
            df_sum_nds.loc[_cm, '淨缺量'] = _cs
        df_sum_nds.loc[_mask, '庫存來源'] = df_sum_nds.loc[_mask, '庫存來源'].apply(
            lambda x: x + '【共用】' if '【共用】' not in str(x) else x)
    shortage_nds = df_sum_nds[(df_sum_nds['淨缺量'] > 0) & (df_sum_nds['庫存來源'] != '不追蹤')].copy()
    _nds_uniq = df_det_nds[['品號','材料品號','生產量']].drop_duplicates(subset=['品號','材料品號'])
    nds_order_boxes = (_nds_uniq[_nds_uniq['材料品號'].isin(shortage_nds['材料品號'])]
                       .groupby('材料品號')['生產量'].sum().round(0).astype(int).to_dict())
    nds_order_count = (_nds_uniq[_nds_uniq['材料品號'].isin(shortage_nds['材料品號'])]
                       .groupby('材料品號')['品號'].nunique().to_dict())
    _nds_pnos = set(df_det_nds['品號'])
    print(f"排除大樹後: {len(_nds_pnos)} 品號  缺料: {len(shortage_nds)} 項")
else:
    df_sum_nds = pd.DataFrame(); shortage_nds = pd.DataFrame()
    nds_order_boxes = {}; nds_order_count = {}; _nds_pnos = set()

tracked  = df_sum[df_sum['庫存來源'] != '不追蹤']
n_print  = (df_sum['庫存來源'] == '需叫布').sum()
print(f"\n彙總: {len(df_sum)} 個品號 (追蹤: {len(tracked)}, 不追蹤包材等: {len(df_sum)-len(tracked)})")
if n_print:
    print(f"  其中印花外層（需叫布）: {n_print} 項")
print(f"缺料: {len(shortage)} 項")
print("\n=== 庫存比對結果（追蹤品項）===")
for _, r in tracked.sort_values('材料類型').iterrows():
    flag = '❌' if r['淨缺量'] > 0 else '✓ '
    print(f"  {flag} {r['材料品號']:32} 需{r['總需備量']:8.1f}{r['單位']}  庫存{r['庫存量']:8.1f}  缺{r['淨缺量']:7.1f}  [{r['庫存來源']}]")

# ─── 7. 輸出 Excel ──────────────────────────────────────────
out_file = unique_path(os.path.join(outdir, f"缺料分析_{today}.xlsx"))
wb_out = openpyxl.Workbook()

# ── 顏色常數（全部加 FF 前綴確保不透明）──
C_SKIN    = "FFE2EFDA"   # 親膚 - 淡綠
C_MELT    = "FFFFF2CC"   # 熔噴 - 淡黃
C_OUTER   = "FFDDEBF7"   # 外層 - 淡藍
C_EAR     = "FFFCE4D6"   # 耳繩 - 淡橘
C_H_NAVY  = "FF1F4E79"   # 標題深藍
C_H_DEEP  = "FF1F3864"   # 更深藍（缺料清單分類/主標題）
C_SUFF_BG = "FFC6EFCE"   # 充足 背景綠
C_SUFF_FG = "FF007F00"   # 充足 字色綠
C_LACK_BG = "FFFFD966"   # 缺料 背景黃
C_LACK_FG = "FFCC0000"   # 缺料 字色紅
C_PEND_BG = "FFFFD966"   # 待確認 背景黃
C_PEND_FG = "FF808000"   # 待確認 字色橄欖
C_RED     = "FFFFCCCC"   # 庫存=0 資料行底色（淡紅）
C_ORANGE  = "FFFFE5CC"   # 有庫存但缺料 資料行底色（淡橘）
C_BOX_BG  = "FFEAF4FB"   # 盒子：淡藍底
C_PRINT_BG= "FFD9EAD3"   # 印花外層：橄欖綠底

TYPE_BG = {'親膚': C_SKIN, '熔噴': C_MELT, '外層': C_OUTER, '耳繩': C_EAR, '盒子': C_BOX_BG}

def sh(cell, bg=C_H_NAVY, font_size=10):
    """表頭樣式：深藍底、白字、粗體、置中"""
    cell.font      = Font(bold=True, name='微軟正黑體', size=font_size, color='FFFFFFFF')
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def sd(cell, bg=None, bold=False, color=None):
    """資料樣式"""
    cell.font      = Font(name='微軟正黑體', size=10, bold=bold,
                          color=color if color else 'FF000000')
    cell.alignment = Alignment(vertical='center')
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)

def set_col_widths(ws, widths, cols):
    for w, c in zip(widths, cols):
        ws.column_dimensions[c].width = w

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

# ══════════════════════════════════════════════════════
# 工作表1: 訂單明細（只顯示 熔噴/親膚/耳繩）
# ══════════════════════════════════════════════════════
ws1 = wb_out.active
ws1.title = "訂單明細"
H1 = ['訂單批號','品號','品名','生產量','材料類型','材料品號','材料品名',
      '需求米數(m)','每捲/kg量','需備量','單位']
for c, h in enumerate(H1, 1):
    sh(ws1.cell(1, c, h))
set_row_height(ws1, 1, 30)

SHOW_TYPES = {'熔噴', '親膚', '耳繩'}
df_det_show = df_det[df_det['材料類型'].isin(SHOW_TYPES)].copy()

prev_batch = None
for r, (_, row) in enumerate(df_det_show.iterrows(), 2):
    d    = row.to_dict()
    same = (d['訂單批號'] == prev_batch)
    prev_batch = d['訂單批號']
    bg   = TYPE_BG.get(d['材料類型'])
    vals = [
        '' if same else d['訂單批號'],
        '' if same else d['品號'],
        '' if same else d['品名'],
        '' if same else d['生產量'],
        d['材料類型'], d['材料品號'], d['材料品名'],
        d['需求米數(m)'], d.get('每捲/kg量', 0), d['需備量'], d['單位']
    ]
    for c, v in enumerate(vals, 1):
        sd(ws1.cell(r, c, v), bg=bg)
    set_row_height(ws1, r, 18)

set_col_widths(ws1, [11,26,20,8,6,26,34,12,10,10,5], 'ABCDEFGHIJK')
print(f"訂單明細（熔噴/親膚/耳繩）: {len(df_det_show)} 筆")

# ══════════════════════════════════════════════════════
# 工作表2: 材料需求彙總
# ══════════════════════════════════════════════════════
ws2 = wb_out.create_sheet("材料需求彙總")
H2 = ['材料類型','材料品號','品名(系統)','材料品名','庫存對應品項',
      '總需求(m)','總需求量','單位','庫存量','單位','淨缺量','狀態']
for c, h in enumerate(H2, 1):
    sh(ws2.cell(1, c, h))
set_row_height(ws2, 1, 30)

for r, (_, row) in enumerate(
        df_sum[df_sum['庫存來源'] != '不追蹤'].sort_values('材料類型').iterrows(), 2):
    bg = TYPE_BG.get(row['材料類型'])
    sys_name2 = invi_name.get(str(row['材料品號']).strip(), '')
    vals = [row['材料類型'], row['材料品號'], sys_name2, row['材料品名'],
            row['庫存來源'],
            round(row['總需求(m)'], 2), round(row['總需備量'], 2), row['單位'],
            row['庫存量'], row['單位'], row['淨缺量'], row['狀態']]
    for c, v in enumerate(vals, 1):
        sd(ws2.cell(r, c, v), bg=bg)
    # 狀態欄（L欄=第12欄）特殊配色
    status = row['狀態']
    sc = ws2.cell(r, 12)
    if status == '✓ 充足':
        sc.font = Font(name='微軟正黑體', size=10, bold=True, color=C_SUFF_FG)
        sc.fill = PatternFill("solid", fgColor=C_SUFF_BG)
    elif status == '⚠ 缺料':
        if row['庫存量'] == 0:
            sc.font = Font(name='微軟正黑體', size=10, bold=True, color=C_LACK_FG)
            sc.fill = PatternFill("solid", fgColor=C_RED)
        else:
            sc.font = Font(name='微軟正黑體', size=10, bold=True, color=C_LACK_FG)
            sc.fill = PatternFill("solid", fgColor=C_LACK_BG)
    elif '待確認' in str(status):
        sc.font = Font(name='微軟正黑體', size=10, color=C_PEND_FG)
        sc.fill = PatternFill("solid", fgColor=C_PEND_BG)
    set_row_height(ws2, r, 19.5)

set_col_widths(ws2, [6,28,30,36,22,12,10,5,8,5,10,8], 'ABCDEFGHIJKL')

# ══════════════════════════════════════════════════════
# 工作表3: 缺料清單
# ══════════════════════════════════════════════════════
ws3 = wb_out.create_sheet("缺料清單")
inv_date = datetime.now().strftime('%Y/%m/%d')
n_order  = len(active)

# 第1行：大標題（深藍底白字，合併 A1:J1）
title_text = f"缺料清單（基準：訂單生產總表 {n_order}筆訂單，庫存日：{inv_date}，共 {len(shortage)} 項缺料）"
t1 = ws3.cell(1, 1, title_text)
t1.font = Font(bold=True, name='微軟正黑體', size=11, color='FFFFFFFF')
t1.fill = PatternFill("solid", fgColor=C_H_DEEP)
t1.alignment = Alignment(horizontal='left', vertical='center')
ws3.merge_cells("A1:L1")
set_row_height(ws3, 1, 24)

# 第2行：欄位標題（深藍底白字）
H3 = ['材料類型','材料品號','品名(系統)','材料品名','庫存對應品項',
      '需備量','單位','現有庫存','缺少量','急迫程度','訂單量(盒)','訂單筆數']
for c, h in enumerate(H3, 1):
    sh(ws3.cell(2, c, h))
set_row_height(ws3, 2, 27.75)

by_type = {}
for _, row in shortage.iterrows():
    by_type.setdefault(row['材料類型'], []).append(row.to_dict())

# 材料類型中文完整名稱
TYPE_LABEL = {'親膚': '親膚布', '熔噴': '熔噴布', '外層': '外層布',
              '耳繩': '耳繩', '印花外層': '印花外層布（需叫布）',
              '盒子': '包裝盒（盒子）', '其他': '其他材料'}
r = 3
for mtype in ['親膚', '熔噴', '外層', '耳繩', '印花外層', '盒子', '其他']:
    items = by_type.get(mtype, [])
    if not items:
        continue
    # 分類標題行（深藍底白字）
    lbl = TYPE_LABEL.get(mtype, mtype)
    grp = ws3.cell(r, 1, f"▌ {lbl}  ({len(items)} 項)")
    grp.font = Font(bold=True, name='微軟正黑體', size=10, color='FFFFFFFF')
    grp.fill = PatternFill("solid", fgColor=C_H_DEEP)
    grp.alignment = Alignment(horizontal='left', vertical='center')
    ws3.merge_cells(f"A{r}:L{r}")
    set_row_height(ws3, r, 20)
    r += 1

    for item in sorted(items, key=lambda x: -x['淨缺量']):
        if item['庫存來源'] == '需叫布':
            urgency = '📋 需叫布'
            bg = C_PRINT_BG
        else:
            urgency = ("🔴 嚴重" if item['庫存量'] == 0
                       else "🟠 高" if item['淨缺量'] > item['總需備量'] * 0.5
                       else "🟡 中")
            bg = C_RED if item['庫存量'] == 0 else C_ORANGE
        sys_name    = invi_name.get(item['材料品號'], '')
        order_boxes = mat_order_boxes.get(item['材料品號'], 0)
        order_cnt   = mat_order_count.get(item['材料品號'], 0)
        vals = [item['材料類型'], item['材料品號'], sys_name,
                item['材料品名'],
                item['庫存來源'], item['總需備量'], item['單位'],
                item['庫存量'], item['淨缺量'], urgency,
                order_boxes, order_cnt]
        for c, v in enumerate(vals, 1):
            sd(ws3.cell(r, c, v), bg=bg)
        set_row_height(ws3, r, 18)
        r += 1

set_col_widths(ws3, [8,30,30,38,22,10,6,10,8,10,10,8], 'ABCDEFGHIJKL')

# ══════════════════════════════════════════════════════
# 工作表4 大樹缺料 / 工作表5 排除大樹缺料 ← 已移除（2026-04-30）
# 大樹訂單一律納入主缺料清單計算，不再獨立分頁
# ══════════════════════════════════════════════════════
# 工作表6 (原4): 訂單總覽（12欄，按材料類型分欄統計）
# ══════════════════════════════════════════════════════
ws4 = wb_out.create_sheet("訂單總覽")
H4 = ['訂單批號','品號','品名','生產量(盒)',
      '親膚布\n品項數','親膚布\n總需捲',
      '熔噴布\n品項數','熔噴布\n總需捲',
      '外層布\n品項數','外層布\n總需捲',
      '耳繩\n品項數', '耳繩\n總需kg']
for c, h in enumerate(H4, 1):
    sh(ws4.cell(1, c, h))
set_row_height(ws4, 1, 31.5)

# 預先建立每品號的材料類型彙總
order_mat = df_det.groupby(['品號','材料類型']).agg(
    品項數=('材料品號','nunique'),
    總需備量=('需備量','sum')
).reset_index()

shortage_pnos = set(df_det[df_det['材料品號'].isin(shortage['材料品號'])]['品號'])

for r, (_, o) in enumerate(active.iterrows(), 2):
    pno  = str(o['品號']).strip()
    grp  = order_mat[order_mat['品號'] == pno]

    def get_type(mtype, col):
        row_t = grp[grp['材料類型'] == mtype]
        if row_t.empty: return 0
        return round(row_t[col].values[0], 2)

    skin_n  = get_type('親膚', '品項數')
    skin_q  = get_type('親膚', '總需備量')
    melt_n  = get_type('熔噴', '品項數')
    melt_q  = get_type('熔噴', '總需備量')
    out_n   = get_type('外層', '品項數')
    out_q   = get_type('外層', '總需備量')
    ear_n   = get_type('耳繩', '品項數')
    ear_q   = get_type('耳繩', '總需備量')

    vals = [o.get('批號',''), pno, o.get('品名',''), o['生產量'],
            skin_n, skin_q, melt_n, melt_q,
            out_n,  out_q,  ear_n,  ear_q]
    for c, v in enumerate(vals, 1):
        cell = ws4.cell(r, c, v)
        # 按欄位設定材料類型底色
        if   c in (5, 6): bg = C_SKIN
        elif c in (7, 8): bg = C_MELT
        elif c in (9,10): bg = C_OUTER
        elif c in (11,12):bg = C_EAR
        else:             bg = None
        sd(cell, bg=bg)
    set_row_height(ws4, r, 18)

set_col_widths(ws4, [11,26,20,9,7,10,7,10,7,10,7,10], 'ABCDEFGHIJKL')

# ══════════════════════════════════════════════════════
# 工作表5: 庫存解析驗證（每日快速確認用）
# ══════════════════════════════════════════════════════
ws5 = wb_out.create_sheet("庫存解析驗證")

# 標題列
ws5.merge_cells('A1:F1')
title_cell = ws5.cell(1, 1, f"庫存解析驗證  （{datetime.now().strftime('%Y/%m/%d')} 執行）— 請掃描確認數量是否正確")
title_cell.font      = Font(bold=True, name='微軟正黑體', size=11, color='FFFFFFFF')
title_cell.fill      = PatternFill("solid", fgColor=C_H_DEEP)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
set_row_height(ws5, 1, 28)

H5 = ['類別', '庫存原始字串（原物料庫存檔）', '解析量', '單位', '備註']
for c, h in enumerate(H5, 1):
    sh(ws5.cell(2, c, h))
set_row_height(ws5, 2, 22)

# 依類別排序輸出
cat_order = {'熔噴': 0, '外層': 1, '親膚': 2}
sorted_inv = sorted(inv_table, key=lambda x: (cat_order.get(x['cat'], 9), x['cat'], x.get('width', 0), x.get('kw', '')))

r5 = 3
prev_cat = None
for it in sorted_inv:
    cat  = it['cat']
    raw  = it['raw']
    qty  = it.get('qty_rolls', it['qty'])
    # 外層/熔噴 顯示捲數，親膚顯示捲數
    unit = '捲'

    # 分類換行標題
    if cat != prev_cat:
        if prev_cat is not None:
            r5 += 1  # 空一行
        ws5.merge_cells(f'A{r5}:F{r5}')
        sec = ws5.cell(r5, 1, f'【{cat}】')
        sec.font      = Font(bold=True, name='微軟正黑體', size=10, color='FFFFFFFF')
        sec.fill      = PatternFill("solid", fgColor=TYPE_BG.get(cat, 'FFDDDDDD').replace('FF', '') if len(TYPE_BG.get(cat, 'FFDDDDDD')) == 10 else TYPE_BG.get(cat, 'FFDDDDDD'))
        sec.fill      = PatternFill("solid", fgColor=TYPE_BG.get(cat, C_H_NAVY))
        sec.alignment = Alignment(vertical='center')
        set_row_height(ws5, r5, 18)
        r5 += 1
        prev_cat = cat

    bg = TYPE_BG.get(cat)
    # 數量為 0 → 淡紅底警示
    row_bg = 'FFFFCCCC' if qty == 0 else bg

    note = ''
    if qty == 0:
        note = '⚠ 庫存為零'

    vals = [cat, raw, qty, unit, note]
    for c, v in enumerate(vals, 1):
        cell = ws5.cell(r5, c, v)
        sd(cell, bg=row_bg)
        if c == 3:  # 數量欄加粗
            cell.font = Font(bold=(qty == 0), name='微軟正黑體', size=10,
                             color='FFCC0000' if qty == 0 else 'FF000000')
    set_row_height(ws5, r5, 18)
    r5 += 1

# 耳繩也列出
if ear_table:
    r5 += 1
    ws5.merge_cells(f'A{r5}:F{r5}')
    sec = ws5.cell(r5, 1, '【耳繩】')
    sec.font      = Font(bold=True, name='微軟正黑體', size=10, color='FF000000')
    sec.fill      = PatternFill("solid", fgColor=C_EAR)
    sec.alignment = Alignment(vertical='center')
    set_row_height(ws5, r5, 18)
    r5 += 1
    for e in ear_table:
        qty_e = e['qty']
        row_bg = 'FFFFCCCC' if qty_e == 0 else C_EAR
        note_e = '⚠ 庫存為零' if qty_e == 0 else ''
        vals = ['耳繩', e['raw'], qty_e, 'kg', note_e]
        for c, v in enumerate(vals, 1):
            cell = ws5.cell(r5, c, v)
            sd(cell, bg=row_bg)
            if c == 3:
                cell.font = Font(bold=(qty_e == 0), name='微軟正黑體', size=10,
                                 color='FFCC0000' if qty_e == 0 else 'FF000000')
        set_row_height(ws5, r5, 18)
        r5 += 1

set_col_widths(ws5, [8, 50, 10, 6, 15], 'ABCDE')
ws5.freeze_panes = 'A3'

# ══════════════════════════════════════════════════════
# 工作表: 須領料（依品項分組，依生產時間排序）
# ══════════════════════════════════════════════════════
ws_pick = wb_out.create_sheet("須領料")

PICK_HEADERS = ['品項', '品號', '品名', '機台', '生產時間', '訂單量', '差異(生產量)', '製令單號', '批號']
PICK_COLS    = ['品項', '品號', '品名', '機台', '生產時間', '訂單量', '差異',         '製令單號', '批號']

# ── 準備資料（用去重複前的完整清單，確保每筆訂單都顯示）──
pick_df = active_for_pick.copy()
for col in PICK_COLS:
    if col not in pick_df.columns:
        pick_df[col] = ''

# 解析生產時間前段日期（用於排序）："4/20~4/30" → 420
def _parse_prod_start(t):
    m = re.match(r'(\d+)[/月](\d+)', str(t).strip())
    if m:
        return int(m.group(1)) * 100 + int(m.group(2))
    return 9999

pick_df['_sort_time'] = pick_df['生產時間'].apply(_parse_prod_start)
# 品項空白者填「未分類」
pick_df['品項'] = pick_df['品項'].fillna('').astype(str).str.strip()
pick_df['品項'] = pick_df['品項'].replace('', '未分類')
pick_df = pick_df.sort_values(['品項', '_sort_time'], na_position='last').reset_index(drop=True)

# 品項共有幾種
all_品項 = pick_df['品項'].unique().tolist()

# ── 大標題 ──
n_pick = len(pick_df)
t0 = ws_pick.cell(1, 1, f"須領料清單（{datetime.now().strftime('%Y/%m/%d')}，共 {n_pick} 筆有效訂單，{len(all_品項)} 個品項）")
t0.font      = Font(bold=True, name='微軟正黑體', size=11, color='FFFFFFFF')
t0.fill      = PatternFill("solid", fgColor=C_H_DEEP)
t0.alignment = Alignment(horizontal='left', vertical='center')
ws_pick.merge_cells("A1:I1")
set_row_height(ws_pick, 1, 26)

# ── 底色設定 ──
C_PICK_GRP  = "FF2C5282"   # 深藍（品項群組標題）
C_PICK_HEAD = "FF3182CE"   # 中藍（欄位標題）
C_PICK_ODD  = "FFEEF6FF"   # 淡藍（奇數資料行）
C_PICK_EVEN = "FFFFFFFF"   # 白（偶數資料行）

r = 2
data_row_in_group = 0   # 群組內資料行計數（交替底色用）

for 品項, grp_data in pick_df.groupby('品項', sort=False):
    # ── 品項群組標題列 ──
    gc = ws_pick.cell(r, 1, f"▌ {品項}　（{len(grp_data)} 筆）")
    gc.font      = Font(bold=True, name='微軟正黑體', size=10, color='FFFFFFFF')
    gc.fill      = PatternFill("solid", fgColor=C_PICK_GRP)
    gc.alignment = Alignment(horizontal='left', vertical='center')
    ws_pick.merge_cells(f"A{r}:I{r}")
    set_row_height(ws_pick, r, 22)
    r += 1

    # ── 欄位標題列 ──
    for c, h in enumerate(PICK_HEADERS, 1):
        hc = ws_pick.cell(r, c, h)
        hc.font      = Font(bold=True, name='微軟正黑體', size=10, color='FFFFFFFF')
        hc.fill      = PatternFill("solid", fgColor=C_PICK_HEAD)
        hc.alignment = Alignment(horizontal='center', vertical='center')
    set_row_height(ws_pick, r, 22)
    r += 1

    data_row_in_group = 0
    for _, row in grp_data.iterrows():
        data_row_in_group += 1
        bg = C_PICK_ODD if data_row_in_group % 2 == 1 else C_PICK_EVEN

        def _v(col):
            v = row.get(col, '')
            if pd.isna(v): return ''
            return v

        # 差異欄：優先用差異，若空則用生產量
        差異_v = _v('差異')
        if 差異_v == '' or 差異_v == 0:
            差異_v = _v('生產量')

        vals = [
            str(_v('品項')),
            str(_v('品號')),
            str(_v('品名')),
            str(_v('機台')),
            str(_v('生產時間')),
            _v('訂單量'),
            差異_v,
            str(_v('製令單號')),
            str(_v('批號')),
        ]
        for c, v in enumerate(vals, 1):
            dc = ws_pick.cell(r, c, v)
            dc.font      = Font(name='微軟正黑體', size=10)
            dc.fill      = PatternFill("solid", fgColor=bg)
            dc.alignment = Alignment(
                horizontal='center' if c in (4, 5, 6, 7) else 'left',
                vertical='center')
        set_row_height(ws_pick, r, 19)
        r += 1

    # 群組間隔空行
    ws_pick.row_dimensions[r].height = 6
    r += 1

set_col_widths(ws_pick, [18, 30, 34, 8, 16, 8, 12, 18, 22], 'ABCDEFGHI')
ws_pick.freeze_panes = 'A2'
print(f"須領料: {n_pick} 筆訂單 | {len(all_品項)} 個品項")

# ── 儲存 ────────────────────────────────────────────────
wb_out.save(out_file)
print(f"\n{'='*55}")
print(f"✅ 完成！")
print(f"   輸出: {out_file}")
print(f"   訂單: {len(active)} 筆 | 材料: {len(df_sum)} 項 | 缺料: {len(shortage)} 項")
os.startfile(outdir)
