# -*- coding: utf-8 -*-
"""
應付帳款自動生成工具
使用方式：執行後選擇兩個檔案，點「開始生成」
-------------------------------------------------
每月流程：
  1. 選擇當月「20XX-MM應付帳款.XLSX」（ERP匯出的單頭資料）
  2. 選擇「應附憑單20XX財務版.xlsx」
  3. 點擊「開始生成」→ 自動新增當月工作表
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import pandas as pd
import copy, os, sys
from datetime import datetime

# ══════════════════════════════════════════════════════════
#  ★ 廠商對照設定（ERP原始名 → 財務版顯示名）
#    若有新廠商，在下方新增一行即可
# ══════════════════════════════════════════════════════════
VENDOR_DISPLAY = {
    '松勝SS':            '松勝SS(外層布.鼻壓條)',
    '彩麗MG':            '彩麗MG(熔噴)',
    '永豐餘':            '永豐餘(紙箱)',
    '創洋股份有限公司TY': '創洋股份有限公司TY',
    '福綿FM':            '福綿FM(親膚)',
    '凱達印刷':          '凱達印刷(紙盒)',
    '喬煒興業':          '喬煒興業(耳繩)',
    '上塬':              '上塬(活性碳)',
    '田一塑膠':          '田一塑膠',
    '綠沅':              '綠沅(耳繩)',
}

# ★ 廠商排列順序（不在清單中的廠商自動排在最後）
VENDOR_ORDER = [
    '松勝SS', '彩麗MG', '永豐餘', '創洋股份有限公司TY',
    '福綿FM', '凱達印刷', '喬煒興業', '上塬', '田一塑膠', '綠沅',
]

# 財務版預設路徑
DEFAULT_FINANCE = r'C:\Users\admin\OneDrive\桌面\應附\應附憑單2026財務版.xlsx'

# ══════════════════════════════════════════════════════════
#  核心邏輯
# ══════════════════════════════════════════════════════════
def generate_ap_sheet(ap_path, finance_path):
    """讀取來源檔，新增月份工作表到財務版"""

    # ── 1. 讀取 單頭資料 ──────────────────────────────────
    df = pd.read_excel(ap_path, sheet_name='單頭資料', header=2)
    df.columns = [str(c).strip() for c in df.columns]

    for col in ['廠商名稱', '發票日期', '發票號碼', '發票貨款', '發票稅額', '發票金額']:
        if col not in df.columns:
            raise ValueError(f'來源檔找不到欄位：「{col}」\n請確認是否選錯檔案。')

    df = df[df['廠商名稱'].notna()].copy()
    df['發票日期'] = pd.to_datetime(df['發票日期'], errors='coerce')
    df['單據日期'] = pd.to_datetime(df.get('單據日期', pd.NaT), errors='coerce')

    # ── 2. 判斷年月 ────────────────────────────────────────
    ref_date = df['單據日期'].dropna()
    if ref_date.empty:
        ref_date = df['發票日期'].dropna()
    year  = int(ref_date.dt.year.mode()[0])
    month = int(ref_date.dt.month.mode()[0])

    roc_year    = year - 1911
    month_label = f'{roc_year}/{month}'
    sheet_title = f'{year}應付帳款明細表({month}月份)'
    sheet_name  = f'應付帳款{year}-{month:02d}'

    # ── 3. 分組整理 ────────────────────────────────────────
    vendors_data = {}
    for _, row in df.iterrows():
        vendor = str(row['廠商名稱']).strip()
        date   = row['發票日期']
        inv_no = str(row['發票號碼']).strip()
        untax  = int(row['發票貨款']) if pd.notna(row['發票貨款']) else 0
        tax    = int(row['發票稅額']) if pd.notna(row['發票稅額']) else 0
        total  = int(row['發票金額']) if pd.notna(row['發票金額']) else 0
        vendors_data.setdefault(vendor, []).append((date, inv_no, untax, tax, total))

    # 依排序清單排列，其餘廠商補在後面
    sorted_vendors = [v for v in VENDOR_ORDER if v in vendors_data]
    sorted_vendors += [v for v in vendors_data if v not in VENDOR_ORDER]

    # ── 4. 開啟財務版，建立新工作表 ────────────────────────
    if not os.path.exists(finance_path):
        raise FileNotFoundError(f'找不到財務版檔案：\n{finance_path}')

    try:
        wb = openpyxl.load_workbook(finance_path)
    except PermissionError:
        raise PermissionError('財務版檔案正在 Excel 中開啟，請先關閉後再執行。')

    if sheet_name in wb.sheetnames:
        raise ValueError(f'工作表「{sheet_name}」已存在！\n請先在 Excel 中刪除後再執行。')

    # 找最後一個「應付帳款」工作表當樣式範本
    template_ws = None
    for sn in reversed(wb.sheetnames):
        if '應付帳款' in sn:
            template_ws = wb[sn]
            break
    if not template_ws:
        raise ValueError('找不到範本工作表（需要至少一個「應付帳款YYYY-MM」工作表）')

    ws = wb.copy_worksheet(template_ws)
    ws.title = sheet_name

    # 清除 R4 以後舊資料
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    ws['A2'].value = sheet_title

    # ── 5. 取範本樣式 ──────────────────────────────────────
    def get_style(src_ws, r, c):
        cell = src_ws.cell(row=r, column=c)
        return {
            'font':          copy.copy(cell.font),
            'fill':          copy.copy(cell.fill),
            'border':        copy.copy(cell.border),
            'alignment':     copy.copy(cell.alignment),
            'number_format': cell.number_format,
        }

    def apply_style(cell, style):
        cell.font          = style['font']
        cell.fill          = style['fill']
        cell.border        = style['border']
        cell.alignment     = style['alignment']
        cell.number_format = style['number_format']

    # 找範本中的資料列、空白列、小計列 row index
    data_ref_row     = 4
    subtotal_ref_row = None
    empty_ref_row    = None
    for r in range(5, 50):
        v = template_ws.cell(row=r, column=1).value
        if v == '小計:':
            subtotal_ref_row = r
            empty_ref_row    = r - 1
            break
    if not subtotal_ref_row:
        subtotal_ref_row = data_ref_row + 2
        empty_ref_row    = data_ref_row + 1

    data_style     = {c: get_style(template_ws, data_ref_row,     c) for c in range(1,8)}
    empty_style    = {c: get_style(template_ws, empty_ref_row,    c) for c in range(1,8)}
    subtotal_style = {c: get_style(template_ws, subtotal_ref_row, c) for c in range(1,8)}

    # ── 6. 填入資料 ────────────────────────────────────────
    cur_row = 4
    for vendor in sorted_vendors:
        invoices     = vendors_data[vendor]
        display_name = VENDOR_DISPLAY.get(vendor, vendor)
        data_start   = cur_row

        for date, inv_no, untax, tax, total in invoices:
            for c in range(1, 8):
                apply_style(ws.cell(row=cur_row, column=c), data_style[c])
            ws.cell(row=cur_row, column=1).value          = month_label
            ws.cell(row=cur_row, column=2).value          = display_name
            ws.cell(row=cur_row, column=3).value          = date
            ws.cell(row=cur_row, column=3).number_format  = 'yyyy/mm/dd'
            ws.cell(row=cur_row, column=4).value          = inv_no
            ws.cell(row=cur_row, column=5).value          = untax
            ws.cell(row=cur_row, column=6).value          = tax
            ws.cell(row=cur_row, column=7).value          = total
            ws.row_dimensions[cur_row].height = 26.25
            cur_row += 1

        data_end = cur_row - 1

        # 空白列
        for c in range(1, 8):
            apply_style(ws.cell(row=cur_row, column=c), empty_style[c])
        ws.row_dimensions[cur_row].height = 26.25
        cur_row += 1

        # 小計列
        for c in range(1, 8):
            apply_style(ws.cell(row=cur_row, column=c), subtotal_style[c])
        ws.cell(row=cur_row, column=1).value = '小計:'
        formula = f'=SUM(G{data_start})' if data_start == data_end else f'=SUM(G{data_start}:G{data_end})'
        ws.cell(row=cur_row, column=7).value = formula
        ws.row_dimensions[cur_row].height = 26.25
        cur_row += 1

    # ── 7. 儲存 ────────────────────────────────────────────
    try:
        wb.save(finance_path)
    except PermissionError:
        raise PermissionError('儲存失敗！財務版檔案可能正在 Excel 中開啟，請關閉後再試。')

    total_amount   = sum(sum(i[4] for i in v) for v in vendors_data.values())
    total_invoices = sum(len(v) for v in vendors_data.values())
    return sheet_name, len(sorted_vendors), total_invoices, total_amount


# ══════════════════════════════════════════════════════════
#  GUI 介面
# ══════════════════════════════════════════════════════════
def main():
    root = tk.Tk()
    root.title('應付帳款自動生成工具')
    root.geometry('560x320')
    root.resizable(False, False)

    # ── 標題 ──
    tk.Label(root, text='應付帳款自動生成工具',
             font=('微軟正黑體', 15, 'bold')).pack(pady=12)

    # ── 來源檔 ──
    frm1 = tk.Frame(root)
    frm1.pack(fill='x', padx=20, pady=4)
    tk.Label(frm1, text='① 來源檔（當月應付帳款）', width=22, anchor='w').pack(side='left')
    ap_var = tk.StringVar()
    tk.Entry(frm1, textvariable=ap_var, width=30).pack(side='left')
    tk.Button(frm1, text='瀏覽', width=6,
              command=lambda: ap_var.set(filedialog.askopenfilename(
                  title='選擇當月應付帳款XLSX',
                  initialdir=r'C:\Users\admin\OneDrive\桌面',
                  filetypes=[('Excel 檔案', '*.xlsx *.XLSX *.xls')]
              ))).pack(side='left', padx=4)

    # ── 財務版 ──
    frm2 = tk.Frame(root)
    frm2.pack(fill='x', padx=20, pady=4)
    tk.Label(frm2, text='② 財務版應付憑單', width=22, anchor='w').pack(side='left')
    fin_var = tk.StringVar(value=DEFAULT_FINANCE)
    tk.Entry(frm2, textvariable=fin_var, width=30).pack(side='left')
    tk.Button(frm2, text='瀏覽', width=6,
              command=lambda: fin_var.set(filedialog.askopenfilename(
                  title='選擇財務版應付憑單',
                  initialdir=r'C:\Users\admin\OneDrive\桌面\應附',
                  filetypes=[('Excel 檔案', '*.xlsx')]
              ))).pack(side='left', padx=4)

    # ── 提示 ──
    tk.Label(root, text='※ 執行前請確認兩個 Excel 檔案已關閉',
             fg='gray', font=('微軟正黑體', 9)).pack()

    # ── 結果顯示 ──
    result_var = tk.StringVar()
    tk.Label(root, textvariable=result_var,
             font=('微軟正黑體', 10),
             wraplength=520, justify='left').pack(pady=6)

    # ── 生成按鈕 ──
    def run():
        result_var.set('處理中，請稍候...')
        root.update()
        ap  = ap_var.get().strip()
        fin = fin_var.get().strip()
        if not ap:
            messagebox.showerror('錯誤', '請選擇來源檔（當月應付帳款XLSX）')
            result_var.set('')
            return
        if not os.path.exists(ap):
            messagebox.showerror('錯誤', f'找不到來源檔：\n{ap}')
            result_var.set('')
            return
        try:
            sheet_name, vendors, invoices, total = generate_ap_sheet(ap, fin)
            msg = (f'✅ 完成！\n'
                   f'新增工作表：{sheet_name}\n'
                   f'廠商 {vendors} 家  ·  發票 {invoices} 張  ·  合計 {total:,} 元')
            result_var.set(msg)
            # 自動開啟財務版
            os.startfile(fin)
        except Exception as e:
            messagebox.showerror('執行錯誤', str(e))
            result_var.set(f'❌ {e}')

    tk.Button(root, text='▶  開始生成',
              font=('微軟正黑體', 12, 'bold'),
              bg='#2E7D32', fg='white',
              padx=24, pady=6,
              command=run).pack(pady=8)

    root.mainloop()


if __name__ == '__main__':
    main()
