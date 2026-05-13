# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

output_path = r"C:\Users\admin\OneDrive\桌面\熔噴布訂購清單_0511.xlsx"

# 現有庫存（來自0511缺料分析）
INV = {
    'BMB-MG-MEDW201751500': 2,
    'BMB-MG-MEDG201751500': 30,
    'BMB-MG-H11W201751500': 6,
    'BMB-MG-MEDG202601500': 8,
    'BMB-MG-H12S252601200': 0,
}

items = [
    {'BMB品號':'BMB-MG-MEDW201751500','規格':'醫白 20×175mm 1500m/捲', '5/8下單':168,'5/11-12到貨':64, '6-7-8月下單':336},
    {'BMB品號':'BMB-MG-MEDG201751500','規格':'醫灰 20×175mm 1500m/捲', '5/8下單':64, '5/11-12到貨':20, '6-7-8月下單':128},
    {'BMB品號':'BMB-MG-H11W201751500','規格':'H11白 20×175mm 1500m/捲','5/8下單':24, '5/11-12到貨':12, '6-7-8月下單':48},
    {'BMB品號':'BMB-MG-MEDG202601500','規格':'醫用灰 20×260mm 1500m/捲','5/8下單':36,'5/11-12到貨':24, '6-7-8月下單':72},
    {'BMB品號':'BMB-MG-H12S252601200','規格':'H12銀 25×260mm 1200m/捲','5/8下單':12, '5/11-12到貨':0,  '6-7-8月下單':24},
]

for d in items:
    bid = d['BMB品號']
    d['現有庫存'] = INV.get(bid, 0)
    d['5月待到貨'] = d['5/8下單'] - d['5/11-12到貨']
    d['6-8月下單'] = d.pop('6-7-8月下單')
    d['總下單量']  = d['5/8下單'] + d['6-8月下單']

total = {
    'BMB品號':'★ 合計','規格':'',
    '現有庫存':    sum(d['現有庫存']    for d in items),
    '5/8下單':     sum(d['5/8下單']     for d in items),
    '5/11-12到貨': sum(d['5/11-12到貨'] for d in items),
    '5月待到貨':   sum(d['5月待到貨']   for d in items),
    '6-8月下單':   sum(d['6-8月下單']   for d in items),
    '總下單量':    sum(d['總下單量']     for d in items),
}

# 欄位排序
col_order = ['BMB品號','規格','現有庫存','5/8下單','5/11-12到貨','5月待到貨','6-8月下單','總下單量']
df = pd.DataFrame(items + [total])[col_order]
df.to_excel(output_path, index=False, sheet_name='熔噴訂購清單')

# ── 美化 ──────────────────────────────────────────────
wb = load_workbook(output_path)
ws = wb.active

thin  = Side(style='thin',   color='BBBBBB')
thick = Side(style='medium', color='333333')
def mk_border(top_thick=False):
    t = thick if top_thick else thin
    return Border(left=thin, right=thin, top=t, bottom=thin)

# 欄寬
for i, w in enumerate([32, 26, 10, 10, 14, 12, 12, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 標題
for cell in ws[1]:
    cell.fill      = PatternFill("solid", fgColor="1F4E79")
    cell.font      = Font(color="FFFFFF", bold=True, size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = mk_border()
ws.row_dimensions[1].height = 32

# 資料列顏色
ROW_COLORS = ["FFFFFF", "D6E4F0"]
for ri, row in enumerate(ws.iter_rows(min_row=2, max_row=len(items)+1)):
    bg = ROW_COLORS[ri % 2]
    for cell in row:
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.font      = Font(size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = mk_border()
    ws.row_dimensions[ri+2].height = 22

# 5月待到貨>0 橘色提示
orange_font = Font(color="C55A11", bold=True, size=11)
for row in ws.iter_rows(min_row=2, max_row=len(items)+1):
    cell = row[5]  # 5月待到貨欄
    if isinstance(cell.value, (int,float)) and cell.value > 0:
        cell.font = orange_font

# 現有庫存=0 紅色
red_font = Font(color="C00000", bold=True, size=11)
for row in ws.iter_rows(min_row=2, max_row=len(items)+1):
    cell = row[2]  # 現有庫存欄
    if isinstance(cell.value, (int,float)) and cell.value == 0:
        cell.font = red_font

# 5/11-12到貨=0 灰色虛線
grey_font = Font(color="AAAAAA", size=11)
for row in ws.iter_rows(min_row=2, max_row=len(items)+1):
    cell = row[4]
    if isinstance(cell.value, (int,float)) and cell.value == 0:
        cell.font  = grey_font
        cell.value = '-'

# 合計列
total_row = ws[len(items)+2]
for cell in total_row:
    cell.fill      = PatternFill("solid", fgColor="FFD966")
    cell.font      = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = mk_border(top_thick=True)
ws.row_dimensions[len(items)+2].height = 26

wb.save(output_path)

# 輸出摘要
print("完成！")
print(f"\n{'規格':<18} {'現貨':>4} {'5/8下單':>7} {'已到貨':>7} {'待到貨':>7} {'6-8月':>7} {'總計':>7}")
print("-"*60)
for d in items:
    print(f"{d['規格']:<18} {d['現有庫存']:>4} {d['5/8下單']:>7} {d['5/11-12到貨']:>7} {d['5月待到貨']:>7} {d['6-8月下單']:>7} {d['總下單量']:>7}")
print("-"*60)
print(f"{'合計':<18} {total['現有庫存']:>4} {total['5/8下單']:>7} {total['5/11-12到貨']:>7} {total['5月待到貨']:>7} {total['6-8月下單']:>7} {total['總下單量']:>7}")
