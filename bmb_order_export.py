# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

output_path = r"C:\Users\admin\OneDrive\桌面\熔噴布訂購清單_0508.xlsx"

# ── 5月下單 ──────────────────────────────────────────────
may_data = [
    {'BMB品號': 'BMB-MG-MEDW201751500', '規格': '醫白 20×175mm 1500m/捲',  '下單量(捲)': 168, '5/11-12首批到貨': 66,  '現有庫存': 10, '備註': '最優先'},
    {'BMB品號': 'BMB-MG-MEDG201751500', '規格': '醫灰 20×175mm 1500m/捲',  '下單量(捲)': 64,  '5/11-12首批到貨': 20,  '現有庫存': 30, '備註': '優先'},
    {'BMB品號': 'BMB-MG-H11W201751500', '規格': 'H11白 20×175mm 1500m/捲', '下單量(捲)': 24,  '5/11-12首批到貨': 10,  '現有庫存': 6,  '備註': ''},
    {'BMB品號': 'BMB-MG-MEDG202601500', '規格': '醫用灰 20×260mm 1500m/捲','下單量(捲)': 36,  '5/11-12首批到貨': 16,  '現有庫存': 8,  '備註': ''},
    {'BMB品號': 'BMB-MG-H12S252601200', '規格': 'H12銀 25×260mm 1200m/捲', '下單量(捲)': 12,  '5/11-12首批到貨': 8,   '現有庫存': 0,  '備註': '庫存歸零'},
]
may_total = {'BMB品號': '5月合計', '規格': '', '下單量(捲)': sum(d['下單量(捲)'] for d in may_data),
             '5/11-12首批到貨': sum(d['5/11-12首批到貨'] for d in may_data),
             '現有庫存': sum(d['現有庫存'] for d in may_data), '備註': ''}

# ── 6-7-8月下單 ──────────────────────────────────────────
jun_data = [
    {'BMB品號': 'BMB-MG-MEDW201751500', '規格': '醫白 20×175mm 1500m/捲',  '下單量(捲)': 336, '5/11-12首批到貨': '', '現有庫存': '', '備註': ''},
    {'BMB品號': 'BMB-MG-MEDG201751500', '規格': '醫灰 20×175mm 1500m/捲',  '下單量(捲)': 128, '5/11-12首批到貨': '', '現有庫存': '', '備註': ''},
    {'BMB品號': 'BMB-MG-H11W201751500', '規格': 'H11白 20×175mm 1500m/捲', '下單量(捲)': 48,  '5/11-12首批到貨': '', '現有庫存': '', '備註': ''},
    {'BMB品號': 'BMB-MG-MEDG202601500', '規格': '醫用灰 20×260mm 1500m/捲','下單量(捲)': 72,  '5/11-12首批到貨': '', '現有庫存': '', '備註': ''},
    {'BMB品號': 'BMB-MG-H12S252601200', '規格': 'H12銀 25×260mm 1200m/捲', '下單量(捲)': 24,  '5/11-12首批到貨': '', '現有庫存': '', '備註': ''},
]
jun_total = {'BMB品號': '6-7-8月合計', '規格': '', '下單量(捲)': sum(d['下單量(捲)'] for d in jun_data),
             '5/11-12首批到貨': '', '現有庫存': '', '備註': ''}

# ── 總合計 ──────────────────────────────────────────────
grand_total = {'BMB品號': '★ 總計', '規格': '',
               '下單量(捲)': sum(d['下單量(捲)'] for d in may_data) + sum(d['下單量(捲)'] for d in jun_data),
               '5/11-12首批到貨': '', '現有庫存': '', '備註': ''}

all_rows = may_data + [may_total] + jun_data + [jun_total] + [grand_total]
df = pd.DataFrame(all_rows)
df.to_excel(output_path, index=False, sheet_name='熔噴訂購清單')

# ── 美化 ──────────────────────────────────────────────
wb = load_workbook(output_path)
ws = wb.active

thin  = Side(style='thin',   color='AAAAAA')
thick = Side(style='medium', color='333333')
border     = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
border_top = Border(left=thin,  right=thin,  top=thick, bottom=thin)

# 欄寬
for i, w in enumerate([30, 28, 14, 18, 12, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 色彩定義
C_HEADER   = "1F4E79"  # 深藍 - 標題
C_MAY      = "D6E4F0"  # 淡藍 - 5月資料
C_JUN      = "E8F5E9"  # 淡綠 - 6-7-8月資料
C_SUBTOTAL = "FFD966"  # 黃   - 小計
C_GRAND    = "F4B942"  # 深黃 - 總計
C_RED      = "C00000"  # 紅   - 重要備註

n_may = len(may_data)       # 5
n_jun = len(jun_data)       # 5
row_may_total  = 1 + n_may + 1          # row 7
row_jun_start  = row_may_total + 1      # row 8
row_jun_total  = row_jun_start + n_jun  # row 13
row_grand      = row_jun_total + 1      # row 14

def style_row(ws, row_idx, fill_color, font_bold=False, font_size=11, top_thick=False):
    fill   = PatternFill("solid", fgColor=fill_color)
    font   = Font(bold=font_bold, size=font_size)
    b      = border_top if top_thick else border
    for cell in ws[row_idx]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = b

# 標題列
style_row(ws, 1, C_HEADER, font_bold=True)
for cell in ws[1]:
    cell.font = Font(color="FFFFFF", bold=True, size=11)

# 5月資料列
for r in range(2, 2 + n_may):
    bg = "FFFFFF" if r % 2 == 0 else C_MAY
    style_row(ws, r, bg)

# 5月小計
style_row(ws, row_may_total, C_SUBTOTAL, font_bold=True, top_thick=True)

# 6-7-8月資料列
for r in range(row_jun_start, row_jun_start + n_jun):
    bg = "FFFFFF" if r % 2 == 0 else C_JUN
    style_row(ws, r, bg, top_thick=(r == row_jun_start))

# 6-7-8月小計
style_row(ws, row_jun_total, C_SUBTOTAL, font_bold=True, top_thick=True)

# 總計
style_row(ws, row_grand, C_GRAND, font_bold=True, font_size=12, top_thick=True)

# 備註紅色
red_font = Font(color=C_RED, bold=True)
for row in ws.iter_rows(min_row=2):
    if row[-1].value in ['最優先', '庫存歸零']:
        row[-1].font = red_font

# 列高
for r in ws.iter_rows():
    ws.row_dimensions[r[0].row].height = 22
ws.row_dimensions[1].height = 28
ws.row_dimensions[row_grand].height = 26

wb.save(output_path)
print(f"已輸出: {output_path}")
