import pandas as pd, shutil, tempfile, os, sys, glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

def sr(path, **kw):
    ext = os.path.splitext(path)[1]
    t = tempfile.mktemp(suffix=ext)
    shutil.copy2(path, t)
    df = pd.read_excel(t, **kw)
    os.unlink(t)
    return df

# ── 庫存 ──
inv = sr(r"C:\Users\admin\OneDrive\桌面\INVR1820260507000012202605070001.xlsx", dtype=str)
inv['庫存包裝數量'] = pd.to_numeric(inv['庫存包裝數量'], errors='coerce').fillna(0)
ly_inv = inv[inv['品號'].fillna('').str.startswith('BEL-LY')].groupby(['品號','品名'])['庫存包裝數量'].sum().reset_index()
ly_inv.columns = ['耳繩品號','耳繩品名','現庫存_箱']

# ── BOM ──
bom_files = glob.glob(r"C:\Users\admin\OneDrive\桌面\CODE資料\BOM*.xlsx")
bom_files.sort(key=os.path.getmtime, reverse=True)
df_bom = sr(bom_files[0], header=None, dtype=str)
df_bom.columns = df_bom.iloc[0]; df_bom = df_bom.iloc[1:].reset_index(drop=True)
c_mat = df_bom.columns[0]; c_prod = df_bom.columns[1]; c_use = df_bom.columns[8]

ly_rows = df_bom[df_bom[c_mat].fillna('').str.startswith('BEL-LY') & df_bom[c_prod].fillna('').ne('')][
    [c_mat, c_prod, c_use]].copy()
ly_rows.columns = ['耳繩品號','成品品號','用量']
ly_rows['用量_m'] = pd.to_numeric(ly_rows['用量'], errors='coerce').fillna(0)

# ── 製令 ──
mfg = sr(r"C:\Users\admin\OneDrive\桌面\CODE資料\製令總表_合併_0428.xlsx", dtype=str)
mfg['開單日期'] = pd.to_datetime(mfg['開單日期'], errors='coerce')
mfg['已生產量'] = pd.to_numeric(mfg['已生產量'], errors='coerce').fillna(0)
mfg_done = mfg[mfg['狀態碼'].isin(['Y:已完工','y:指定完工']) & (mfg['已生產量'] > 0)].copy()
mfg_done['產品品號'] = mfg_done['產品品號'].str.strip()

merged = mfg_done.merge(ly_rows, left_on='產品品號', right_on='成品品號', how='inner')
M_PER_BOX = 2.777 * 730  # ≈2027m/箱

# 統計
grp = merged.groupby('耳繩品號')
stats = grp.agg(製令次數=('製令單號','count'), 總生產量_盒=('已生產量','sum'), 最近使用=('開單日期','max')).reset_index()
耗用 = grp.apply(lambda g: (g['用量_m'] * g['已生產量']).sum() / M_PER_BOX).reset_index()
耗用.columns = ['耳繩品號','歷史耗用_箱']
stats = stats.merge(耗用, on='耳繩品號')
stats['歷史耗用_箱'] = stats['歷史耗用_箱'].round(1)
stats['最近使用'] = stats['最近使用'].dt.strftime('%Y-%m')

# 合併庫存
result = stats.merge(ly_inv, on='耳繩品號', how='outer')
result['製令次數'] = result['製令次數'].fillna(0).astype(int)
result['歷史耗用_箱'] = result['歷史耗用_箱'].fillna(0)
result['現庫存_箱'] = result['現庫存_箱'].fillna(0)
result['最近使用'] = result['最近使用'].fillna('—')
result = result.sort_values('製令次數', ascending=False).reset_index(drop=True)

# ── 建議補購邏輯 ──
# 已有342(粉紅342) 14箱，還需湊 26箱
# 優先：高頻 + 庫存偏低（< 月均耗用×2）的品項
ALREADY = 14   # 342已在綠沅
TARGET  = 40
NEED    = TARGET - ALREADY  # 26箱

# 月均耗用（22個月）≈ 歷史耗用/22
result['月均耗用_箱'] = (result['歷史耗用_箱'] / 22).round(1)
# 庫存月數 = 現庫存 / 月均耗用
result['庫存月數'] = result.apply(
    lambda r: round(r['現庫存_箱'] / r['月均耗用_箱'], 1) if r['月均耗用_箱'] > 0 else 99, axis=1)

# 建議補購：頻率前10 中，庫存月數 < 3 的優先補；按缺口大小分配
# 篩出高頻活躍品（製令次數≥10，最近使用≥2026）
active = result[
    (result['製令次數'] >= 10) &
    (result['最近使用'] >= '2026')
].copy()

# 缺口分數：月均耗用越大 / 庫存越少 → 越需要補
active['缺口分'] = (active['月均耗用_箱'] / (active['現庫存_箱'] + 0.5)).round(2)
active = active.sort_values('缺口分', ascending=False)

# 分配建議箱數（整箱，合計=26）
suggest = []
remain = NEED
for _, r in active.iterrows():
    if remain <= 0: break
    # 建議補到 3個月用量，但單色不超過8箱
    ideal = max(0, round(r['月均耗用_箱'] * 3 - r['現庫存_箱']))
    alloc = min(ideal, 8, remain)
    alloc = max(alloc, 1)  # 至少1箱
    alloc = min(alloc, remain)
    suggest.append((r['耳繩品號'], alloc))
    remain -= alloc

# 若還有剩，分給排名最高的
if remain > 0:
    for i in range(len(suggest)):
        if remain <= 0: break
        suggest[i] = (suggest[i][0], suggest[i][1] + 1)
        remain -= 1

sug_dict = dict(suggest)
result['建議補購_箱'] = result['耳繩品號'].map(sug_dict).fillna(0).astype(int)
result.loc[result['耳繩品號'] == 'BEL-LY060-PNK342', '建議補購_箱'] = 0  # 342已有14箱

total_sug = result['建議補購_箱'].sum()

# ── Excel ──
DARK   = '1C3F6B'
GREEN  = '1A5C3A'
YELLOW = 'CC7700'
RED    = 'CC2200'
ROW_A  = 'D6E8FB'; ROW_B = 'EEF5FF'
HLT    = 'FFF0D0'

def hd(ws, r, c, v, bg=DARK):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name='微軟正黑體', bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(bottom=Side(style='medium', color='888888'), right=Side(style='thin', color='BBBBBB'))

def dc(ws, r, c, v, bg='FFFFFF', bold=False, align='left', fmt=None, color='111111', sz=11):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name='微軟正黑體', bold=bold, size=sz, color=color)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center')
    cell.border = Border(right=Side(style='hair', color='DDDDDD'), bottom=Side(style='hair', color='DDDDDD'))
    if fmt: cell.number_format = fmt

wb = Workbook()
ws = wb.active
ws.title = '綠沅耳繩湊箱分析'

# 標題
ws.merge_cells('A1:L1')
t = ws.cell(row=1, column=1, value=f'綠沅6.0耳繩 回台湊箱分析  |  目標: 40箱一版  |  已確認: 新粉紅342 × 14箱  |  還需: {NEED}箱')
t.font = Font(name='微軟正黑體', bold=True, size=14, color='FFFFFF')
t.fill = PatternFill('solid', start_color=DARK)
t.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 32

# 說明列
ws.merge_cells('A2:L2')
n = ws.cell(row=2, column=1, value='※ 依製令歷史頻率排序 | 建議補購優先考慮「高頻 + 庫存月數偏低」品項 | 黃底=建議補購 | 製令資料: 2024-01~2026-04（共22個月）')
n.font = Font(name='微軟正黑體', size=10, color='555555', italic=True)
n.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[2].height = 18

# 表頭
hdrs = ['耳繩品號','耳繩品名','製令次數','歷史耗用(箱)','月均耗用(箱)','現庫存(箱)','庫存月數','缺口評分','在綠沅(箱)','建議補購(箱)','累計箱數','備註']
for ci, h in enumerate(hdrs, 1):
    bg = GREEN if ci == 10 else DARK
    hd(ws, 3, ci, h, bg)
ws.row_dimensions[3].height = 26

running = 0
for ri, (_, row) in enumerate(result.iterrows(), 4):
    sug = int(row['建議補購_箱'])
    is342 = row['耳繩品號'] == 'BEL-LY060-PNK342'
    is_sug = sug > 0

    if is342:
        running += ALREADY
    elif is_sug:
        running += sug

    bg = HLT if is_sug or is342 else (ROW_A if ri % 2 == 0 else ROW_B)

    # 庫存月數顏色
    km = row['庫存月數']
    km_color = RED if km < 2 else (YELLOW if km < 4 else '1A5C3A')

    # 缺口分（只對高頻顯示）
    gap = row.get('缺口分', 0)

    dc(ws, ri, 1, row['耳繩品號'], bg)
    dc(ws, ri, 2, row['耳繩品名'], bg, bold=is_sug or is342)
    dc(ws, ri, 3, row['製令次數'], bg, align='center', bold=row['製令次數']>=50, color='1C3F6B' if row['製令次數']>=50 else '111111')
    dc(ws, ri, 4, row['歷史耗用_箱'], bg, align='right', fmt='#,##0.0')
    dc(ws, ri, 5, row['月均耗用_箱'], bg, align='right', fmt='0.0')
    dc(ws, ri, 6, row['現庫存_箱'], bg, align='right', fmt='#,##0.0')
    dc(ws, ri, 7, km if km < 99 else '—', bg, align='right', color=km_color, bold=(km < 2), fmt='0.0' if km < 99 else None)
    dc(ws, ri, 8, gap if gap > 0 else '', bg, align='right', fmt='0.00' if gap > 0 else None)
    dc(ws, ri, 9, ALREADY if is342 else '', bg, align='center', bold=True, color='CC2200' if is342 else '111111')
    dc(ws, ri, 10, sug if sug > 0 else '', bg, align='center', bold=True, color=GREEN if sug > 0 else '111111', sz=12 if sug > 0 else 11)
    dc(ws, ri, 11, running if (is_sug or is342) else '', bg, align='center', bold=True, color=DARK)

    remark = ''
    if is342: remark = '✅ 已在綠沅 14箱'
    elif is_sug: remark = f'▶ 建議加入 {sug}箱'
    dc(ws, ri, 12, remark, bg, bold=is_sug or is342, color=GREEN if is_sug else ('CC2200' if is342 else '111111'))
    ws.row_dimensions[ri].height = 20

# 合計行
last = len(result) + 4
ws.merge_cells(f'A{last}:I{last}')
sc = ws.cell(row=last, column=1, value=f'建議補購合計: {total_sug} 箱  +  新粉紅342: 14 箱  =  共 {total_sug + ALREADY} 箱')
sc.font = Font(name='微軟正黑體', bold=True, size=12, color='FFFFFF')
sc.fill = PatternFill('solid', start_color=GREEN)
sc.alignment = Alignment(horizontal='center', vertical='center')
ws.cell(row=last, column=10, value=total_sug).font = Font(name='微軟正黑體', bold=True, size=13, color='FFFFFF')
ws.cell(row=last, column=10).fill = PatternFill('solid', start_color=GREEN)
ws.cell(row=last, column=10).alignment = Alignment(horizontal='center', vertical='center')
ws.merge_cells(f'K{last}:L{last}')
ws.cell(row=last, column=11, value=total_sug + ALREADY).font = Font(name='微軟正黑體', bold=True, size=14, color='FFFFFF')
ws.cell(row=last, column=11).fill = PatternFill('solid', start_color=DARK)
ws.cell(row=last, column=11).alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[last].height = 28

# 欄寬
for col, w in zip('ABCDEFGHIJKL', [28, 26, 8, 12, 12, 10, 10, 10, 10, 12, 10, 18]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A4'

out = r"C:\Users\admin\OneDrive\桌面\綠沅耳繩湊箱分析.xlsx"
wb.save(out)
print(f"OK: {out}")
print(f"\n建議補購 {total_sug} 箱（+342已有14箱 = {total_sug+ALREADY}箱）")
for code, qty in sorted(sug_dict.items(), key=lambda x: -x[1]):
    row = result[result['耳繩品號']==code].iloc[0]
    print(f"  {row['耳繩品名']:28} {qty}箱  (庫存{row['現庫存_箱']:.0f}箱, 月均{row['月均耗用_箱']:.1f}箱, 頻率{row['製令次數']}次)")
