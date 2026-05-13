# -*- coding: utf-8 -*-
import pandas as pd, warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

output_path = r"C:\Users\admin\OneDrive\桌面\親膚布訂購清單_0508.xlsx"

# 品名對照
NAMES = {
    'BES-FM-W01-201753000': '親膚布-福綿175-白20gsm',
    'BES-FM-W01-202003000': '親膚布-福綿200-白20gsm',
    'BES-FM-W01-202503000': '親膚布-福綿250-白20gsm',
    'BES-FM-W01-202603000': '親膚布-福綿260-白20gsm',
    'BES-FM-W01-222503000': '親膚布-福綿250-白22gsm',
    'BES-FM-W02-201753000': '親膚布-福綿175-白20gsm(W02)',
    'BES-FM-W02-252602000': '親膚布-福綿260-白25gsm',
    'BES-FM-W03-202603000': '親膚布-福綿260-白20gsm(雙壓)',
    'BES-JG-W01-201753000': '親膚布-嘉谷175-白20gsm',
    'BES-JG-W01-202003000': '親膚布-嘉谷200-白20gsm',
    'BES-JG-W01-202203000': '親膚布-嘉谷220-白20gsm',
    'BES-JG-W01-202603000': '親膚布-嘉谷260-白20gsm',
}

# 最新庫存（來自缺料分析0511）
INV_0511 = {
    'BES-FM-W01-201753000': 6,
    'BES-FM-W01-202003000': 66,   # 更新：72→66
    'BES-FM-W01-202503000': 32,
    'BES-FM-W01-202603000': 32,
    'BES-FM-W01-222503000': 0,
    'BES-FM-W02-201753000': 0,
    'BES-FM-W02-252602000': 17,
    'BES-FM-W03-202603000': 0,
    'BES-JG-W01-201753000': 48,
    'BES-JG-W01-202003000': 18,
    'BES-JG-W01-202203000': 0,
    'BES-JG-W01-202603000': 8,
}

# 月均（來自進貨歷史+實際領出）
AVG = {
    'BES-FM-W01-201753000': 45.9,
    'BES-FM-W01-202003000': 66.4,
    'BES-FM-W01-202503000': 0.0,
    'BES-FM-W01-202603000': 16.9,
    'BES-FM-W01-222503000': 0.8,
    'BES-FM-W02-201753000': 8.4,
    'BES-FM-W02-252602000': 12.2,
    'BES-FM-W03-202603000': 5.5,
    'BES-JG-W01-201753000': 36.3,
    'BES-JG-W01-202003000': 28.1,
    'BES-JG-W01-202203000': 5.2,
    'BES-JG-W01-202603000': 13.0,
}

rows = []
for bid in INV_0511:
    stock = INV_0511[bid]
    avg   = AVG.get(bid, 0)
    may_need = round(avg, 1)
    may_gap  = round(max(0, may_need - stock), 1)
    stock_after = max(0, stock - may_need)
    jun_need = round(avg, 1)
    jun_gap  = round(max(0, jun_need - stock_after), 1)
    total_gap = round(may_gap + jun_gap, 1)
    rows.append({
        'BES品號':    bid,
        '品名':       NAMES.get(bid, ''),
        '現有庫存(捲)': stock,
        '月均(捲)':   avg,
        '5月需求':    may_need,
        '5月缺':      may_gap,
        '6月需求':    jun_need,
        '6月缺':      jun_gap,
        '總缺':       total_gap,
    })

df = pd.DataFrame(rows)
df.to_excel(output_path, index=False, sheet_name='親膚布5-6月預估')

# ── 美化 ──────────────────────────────────────────────
wb = load_workbook(output_path)
ws = wb.active

thin  = Side(style='thin',   color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 欄寬
for i, w in enumerate([32, 30, 14, 10, 10, 10, 10, 10, 10], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 標題
for cell in ws[1]:
    cell.fill      = PatternFill("solid", fgColor="1F4E79")
    cell.font      = Font(color="FFFFFF", bold=True, size=11)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = border

# 內容
red_font = Font(color="C00000", bold=True)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border
        if cell.row % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="E8F5E9")
    # 總缺欄若>0 紅色
    gap_cell = row[8]
    if isinstance(gap_cell.value, (int, float)) and gap_cell.value > 0:
        gap_cell.font = red_font

# 列高
for r in ws.iter_rows():
    ws.row_dimensions[r[0].row].height = 22
ws.row_dimensions[1].height = 28

wb.save(output_path)
print(f"完成: {output_path}")
