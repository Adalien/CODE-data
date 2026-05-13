# -*- coding: utf-8 -*-
import pandas as pd, warnings, sys, io
warnings.filterwarnings('ignore')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

path = r"C:\Users\admin\OneDrive\桌面\20240102-20260508進貨單價.xlsx"
df = pd.read_excel(path)
df.columns = [c.strip() for c in df.columns]

df['品號']     = df['品    號'].astype(str).str.strip()
df['品名']     = df['品    名'].astype(str).str.strip()
df['日期']     = pd.to_datetime(df['單據日期'], errors='coerce')
df['年月']     = df['日期'].dt.to_period('M')
df['進貨捲數'] = pd.to_numeric(df['驗收包裝數量'], errors='coerce').fillna(0)

bes = df[df['品號'].str.startswith('BES')].copy()
recent = bes[bes['日期'] >= pd.Timestamp('2025-05-01')].copy()

monthly = recent.groupby(['年月','品號'])['進貨捲數'].sum().unstack(fill_value=0)
avg = monthly.mean()
avg = avg[avg > 0].sort_values(ascending=False)

BES_INV = {
    'BES-FM-W01-201753000': 6.0,
    'BES-FM-W01-202003000': 72.0,
    'BES-FM-W01-202503000': 32.0,
    'BES-FM-W01-202603000': 32.0,
    'BES-FM-W02-252602000': 17.0,
    'BES-FM-W02-201753000': 0.0,
    'BES-FM-W03-202603000': 0.0,
    'BES-JG-W01-201753000': 48.0,
    'BES-JG-W01-202003000': 18.0,
    'BES-JG-W01-202203000': 0.0,
    'BES-JG-W01-202603000': 8.0,
    'BES-FM-W01-222503000': 0.0,
}

print("="*80)
print("親膚布 5-6月需求預估（依歷史月均進貨量）")
print("="*80)
print(f"\n  {'BES品號':<38} {'現貨':>6} {'5月需求':>8} {'5月缺':>8} {'6月需求':>8} {'6月缺':>8} {'總缺':>8}")
print(f"  {'-'*38} {'-'*6} {'-'*8} {'-'*8} {'-'*8} {'-'*8} {'-'*8}")

t5=0; t6=0; tgap5=0; tgap6=0; tgap=0
all_bids = sorted(set(list(avg.index) + list(BES_INV.keys())))

rows = []
for bid in all_bids:
    a     = round(avg.get(bid, 0), 1)
    stock = BES_INV.get(bid, 0)
    if a == 0 and stock == 0:
        continue
    may_need  = a
    may_gap   = max(0, may_need - stock)
    # 6月庫存 = 現貨用完5月後剩餘（若5月缺就是0）
    stock_after_may = max(0, stock - may_need)
    jun_need  = a
    jun_gap   = max(0, jun_need - stock_after_may)
    total_gap = may_gap + jun_gap

    flag = "❌" if total_gap > 0 else "✓ "
    print(f"  {flag} {bid:<38} {stock:>6.1f} {may_need:>8.1f} {may_gap:>8.1f} {jun_need:>8.1f} {jun_gap:>8.1f} {total_gap:>8.1f}")
    t5+=may_need; t6+=jun_need; tgap5+=may_gap; tgap6+=jun_gap; tgap+=total_gap
    rows.append({'BES品號':bid,'現貨':stock,'5月需求':may_need,'5月缺':may_gap,
                 '6月需求':jun_need,'6月缺':jun_gap,'總缺':total_gap})

print(f"\n  {'合計':<40} {sum(BES_INV.values()):>6.1f} {t5:>8.1f} {tgap5:>8.1f} {t6:>8.1f} {tgap6:>8.1f} {tgap:>8.1f}")
print(f"\n  5月預估需求: {t5:.1f} 捲  缺: {tgap5:.1f} 捲")
print(f"  6月預估需求: {t6:.1f} 捲  缺: {tgap6:.1f} 捲")
print(f"  ★ 總缺口:    {tgap:.1f} 捲")

# ── 匯出 Excel ──────────────────────────────────────────
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

out = r"C:\Users\admin\OneDrive\桌面\親膚布訂購清單_0508.xlsx"
result_df = pd.DataFrame(rows)
result_df.to_excel(out, index=False, sheet_name='親膚布5-6月預估')

wb = load_workbook(out)
ws = wb.active

# 標題
hfill = PatternFill("solid", fgColor="1F4E79")
hfont = Font(color="FFFFFF", bold=True, size=11)
for cell in ws[1]:
    cell.fill = hfill; cell.font = hfont
    cell.alignment = Alignment(horizontal='center', vertical='center')

# 內容
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if cell.row % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="DCE6F1")

# 缺口欄位紅色（5月缺=col3, 6月缺=col5, 總缺=col6, 0-based）
red = Font(color="C00000", bold=True)
for row in ws.iter_rows(min_row=2):
    for ci in [2, 4, 6]:
        if ci < len(row):
            cell = row[ci]
            if isinstance(cell.value, (int,float)) and cell.value > 0:
                cell.font = red

# 欄寬
for i, w in enumerate([38,8,10,10,10,10,10], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].height = 22
ws.row_dimensions[1].height = 28

thin = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for row in ws.iter_rows():
    for cell in row:
        cell.border = border

wb.save(out)
print(f"\n已輸出: {out}")
